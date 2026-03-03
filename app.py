
def _clean_trz(s):
    if not s:
        return ""
    s = str(s).strip()
    # quita backslashes que vienen escapando comillas
    s = s.replace("\\'", "").replace('\\"', "").replace("\\", "")
    # quita comillas envolventes repetidas
    while len(s) >= 2 and ((s[0] == "'" and s[-1] == "'") or (s[0] == '"' and s[-1] == '"')):
        s = s[1:-1].strip()
    # quita comillas sueltas en extremos
    s = s.strip().strip("'").strip('"').strip()
    return s

import re
import time
import uuid

def _last5_digits_from_dm(dm: str) -> str:
    """Extrae los últimos 5 dígitos del DataMatrix. Si hay menos de 5 dígitos, usa los últimos 5 caracteres."""
    s = (dm or "").strip()
    digits = re.findall(r"\d", s)
    if len(digits) >= 5:
        return "".join(digits[-5:])
    s2 = re.sub(r"\s+", "", s)
    return (s2[-5:] if len(s2) >= 5 else s2)

def _build_nombre_trazabilidad(prefijo_nombre: str, codigo_datamatrix: str) -> str:
    pref = (prefijo_nombre or "").strip()
    suf = _last5_digits_from_dm(codigo_datamatrix)
    if not pref and not suf:
        return ""
    return f"{pref}{suf}"

import os
import re
import time
import io
import csv
import json
import base64
from datetime import datetime
from typing import List, Optional
from dotenv import load_dotenv

# Carga variables de entorno desde .env (útil para desarrollo local)
load_dotenv()

# Etiquetas (PDF) + código de barras / QR
from reportlab.pdfgen import canvas
from reportlab.graphics.barcode import code128, qr
from reportlab.lib.units import mm

# Para autocompletar artículos (Articulos.xlsx)
try:
    import pandas as pd
except Exception:  # pragma: no cover
    pd = None

try:
    from openpyxl import load_workbook
except Exception:  # pragma: no cover
    load_workbook = None

from fastapi import FastAPI, Request, UploadFile, File, Form, Depends, HTTPException
from fastapi.responses import FileResponse, HTMLResponse, RedirectResponse, JSONResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
import sys
from pathlib import Path

# Rutas compatibles con PyInstaller (sys._MEIPASS) y ejecución normal
BASE_DIR = Path(getattr(sys, "_MEIPASS", Path(__file__).resolve().parent))

from db import init_db, get_conn
from excel_import import leer_excel_envio

from auth_utils import verify_password, hash_password, make_serializer, sign_session
from security import get_current_user, require_roles

try:
    from openpyxl import Workbook
except Exception:  # pragma: no cover
    Workbook = None

app = FastAPI(title="Xurgical SAT")

# Arranque de la aplicación
print(">>> XURGICAL SAT STARTING UP...", flush=True)
# La inicialización se realiza en el evento on_startup

app.state.secret_key = os.environ.get("XURGICAL_SECRET_KEY", "dev-secret-change-me")
app.state.serializer = make_serializer(app.state.secret_key)

UPLOAD_DIR = os.environ.get("XURGICAL_UPLOAD_DIR", "uploads")
# Priorizamos que las fotos estén dentro del disco persistente si estamos en Render
FOTOS_DIR = os.environ.get("XURGICAL_FOTOS_DIR", os.path.join(UPLOAD_DIR, "fotos"))

try:
    os.makedirs(UPLOAD_DIR, exist_ok=True)
    os.makedirs(FOTOS_DIR, exist_ok=True)
    print(f"DEBUG: Directories created/verified. UPLOAD_DIR={UPLOAD_DIR}", flush=True)
except Exception as e:
    print(f">>> CRITICAL ERROR creating directories: {e}", flush=True)

# Montaje de estáticos
try:
    print(">>> Mounting static files...", flush=True)
    app.mount("/static/fotos", StaticFiles(directory=FOTOS_DIR), name="fotos_externas")
    app.mount("/static", StaticFiles(directory=str(BASE_DIR / "static")), name="static")
    print(">>> Static files mounted.", flush=True)
except Exception as e:
    print(f">>> CRITICAL ERROR mounting static files: {e}", flush=True)

templates = Jinja2Templates(directory=str(BASE_DIR / "templates"))
print(">>> Templates initialized.", flush=True)

def format_fecha(value):
    if not value or value == "-":
        return "-"
    try:
        # Detectar si es un string y tratar de parsearlo
        s = str(value).strip()
        if not s: return "-"
        
        # Intentar varios formatos comunes
        dt = None
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
            try:
                dt = datetime.strptime(s, fmt)
                break
            except:
                continue
        
        if not dt:
            try:
                dt = datetime.fromisoformat(s.replace("Z", "+00:00"))
            except:
                return s # Si falla todo, devuelve el original
        
        return dt.strftime("%d-%m-%Y")
    except:
        return value

templates.env.filters["fecha"] = format_fecha



# -----------------------------
# CLIENTES (prefijo + contador)
# -----------------------------
def _list_clientes(cur) -> list[dict]:
    """Listado de clientes.

    Nota: en instalaciones existentes puede faltar la tabla por no haberse ejecutado
    aún init_db() sobre una BD antigua. Recuperamos automáticamente.
    """
    # Detectar si estamos en Postgres o SQLite para capturar errores de tabla inexistente
    import sqlite3
    try:
        import psycopg2
        PG_ERR = psycopg2.Error
    except ImportError:
        PG_ERR = Exception

    sql = "SELECT id, nombre, prefijo, prefijo_nombre, email, ultimo_numero FROM clientes ORDER BY LOWER(nombre) ASC"
    try:
        cur.execute(sql)
        return [dict(r) for r in cur.fetchall()]
    except (sqlite3.OperationalError, PG_ERR) as e:
        err_msg = str(e).lower()
        if "no such table" in err_msg or "does not exist" in err_msg:
            init_db()
            cur.execute(sql)
            return [dict(r) for r in cur.fetchall()]
        raise


def _envios_has_column(cur, col: str) -> bool:
    try:
        from db import get_table_columns
        cols = get_table_columns(cur, "envios")
        return col in cols
    except Exception:
        return False


def _get_cliente(cur, cliente_id: int) -> dict | None:
    cur.execute(
        "SELECT id, nombre, prefijo, prefijo_nombre, email, ultimo_numero FROM clientes WHERE id=?",
        (int(cliente_id),),
    )
    r = cur.fetchone()
    return dict(r) if r else None


def _reserve_numeros_cliente(cur, cliente_id: int, cantidad: int) -> tuple[str, str, list[int]]:
    """Reserva un rango de numeros (por cliente) de forma transaccional.

    Devuelve: (prefijo_dm, prefijo_nombre, [n1, n2, ...])
    """
    if cantidad <= 0:
        return "", []

    # Bloqueo transaccional (en sqlite3, basta con ejecutar dentro del mismo conn
    # con BEGIN IMMEDIATE para evitar carreras).
    cur.execute("BEGIN IMMEDIATE")

    cur.execute("SELECT prefijo, prefijo_nombre, ultimo_numero FROM clientes WHERE id=?", (int(cliente_id),))
    row = cur.fetchone()
    if not row:
        raise HTTPException(status_code=400, detail="Cliente no encontrado")

    row = dict(row)

    prefijo_dm = (row["prefijo"] or "").strip()
    prefijo_nombre = (row.get("prefijo_nombre") or "").strip()
    ultimo = int(row["ultimo_numero"] or 0)
    nums = list(range(ultimo + 1, ultimo + cantidad + 1))
    cur.execute("UPDATE clientes SET ultimo_numero=? WHERE id=?", (nums[-1], int(cliente_id)))

    return prefijo_dm, prefijo_nombre, nums

# moved to top


# -----------------------------
# ARTICULOS (catálogo para autocompletar)
# -----------------------------
ARTICULOS_XLSX = str(BASE_DIR / 'Articulos.xlsx')
ARTICULOS_XLS  = str(BASE_DIR / 'Articulos.xls')
_articulos_map = None  # dict normalizado: codigo -> {descripcion, fabricante?}

def _norm_codigo(x: str) -> str:
    """Normaliza códigos de artículo.

    Reglas:
    - Trim + upper (búsqueda insensible a mayúsculas)
    - Elimina prefijos del catálogo (p.ej. RP, MT), ignorando mayúsculas/minúsculas para que el usuario pueda
      buscar sin ellos y el sistema siempre devuelva el código "limpio".
    """
    s = (x or '').strip().upper()
    # En el Excel algunos códigos vienen con prefijos como RP/MT.
    # Los quitamos para unificar y evitar que el usuario tenga que escribirlos.
    for pfx in ("RP", "MT"):
        if s.startswith(pfx):
            s = s[len(pfx):]
            break
    # Quita separadores habituales después del prefijo (espacio, guión, etc.)
    s = s.lstrip(" -_\t")
    return s


def _codigo_variants(codigo_norm: str) -> list:
    """Genera variantes de búsqueda para un código normalizado.

    En el catálogo hay códigos que son equivalentes con o sin una "R" final.
    Esta función devuelve una lista de claves candidatas para encontrar el artículo.
    """
    c = (codigo_norm or '').strip().upper()
    if not c:
        return []

    out = []
    def add(x: str):
        x = (x or '').strip().upper()
        if x and x not in out:
            out.append(x)

    add(c)

    # Variante sin espacios internos
    add(c.replace(' ', ''))

    # Equivalencia con/sin R final
    if c.endswith('R') and len(c) > 1:
        add(c[:-1])
        add(c[:-1].replace(' ', ''))
    else:
        add(c + 'R')
        add((c + 'R').replace(' ', ''))

    return out

def load_articulos_map() -> dict:
    """Carga el catálogo de Articulos para autocompletar."""
    global _articulos_map
    if _articulos_map is not None:
        return _articulos_map

    path = ARTICULOS_XLS if os.path.exists(ARTICULOS_XLS) else ARTICULOS_XLSX
    if not os.path.exists(path):
        print(f">>> [ARTICULOS] Archivo no encontrado: {path}", flush=True)
        _articulos_map = {}
        return _articulos_map

    def _put(m, code, desc, fab=None):
        code = _norm_codigo(str(code))
        if not code or code == "NAN":
            return
        desc = ("" if desc is None else str(desc)).strip()
        if not desc or desc.lower() == "nan":
            return
        fab_s = None
        if fab is not None:
            fab_s = str(fab).strip()
            if not fab_s or fab_s.lower() == "nan":
                fab_s = None

        if fab_s is None:
            try:
                d = desc.strip()
                if d.upper().endswith(" AS"):
                    fab_s = "Aescula"
            except Exception:
                pass
        m[code] = {"descripcion": desc, "fabricante": fab_s}
        if code.endswith('R') and len(code) > 1:
            code2 = code[:-1].strip()
            if code2 and code2 not in m:
                m[code2] = {"descripcion": desc, "fabricante": fab_s}

    m = {}
    df = None
    if pd is not None:
        try:
            engine = 'openpyxl' if path.lower().endswith('.xlsx') else None
            df = pd.read_excel(path, engine=engine)
            print(f">>> [ARTICULOS] Cargado {path} ({len(df)} filas)", flush=True)
        except Exception as e:
            print(f">>> [ARTICULOS] Error {path}: {e}", flush=True)
            if path.lower().endswith('.xls') and os.path.exists(ARTICULOS_XLSX):
                try:
                    path = ARTICULOS_XLSX
                    df = pd.read_excel(path, engine='openpyxl')
                except Exception:
                    pass

    if df is not None:
        cols_raw = [str(c).strip() for c in df.columns]
        cols_map = {c.lower(): c for c in cols_raw}
        col_codigo = cols_map.get("código") or cols_map.get("codigo") or cols_map.get("cod") or (cols_raw[0] if cols_raw else None)
        col_desc = cols_map.get("descripción") or cols_map.get("descripcion") or cols_map.get("denominacion") or (cols_raw[1] if len(cols_raw) >= 2 else None)
        col_fab = None
        if len(cols_raw) >= 4:
            col_fab = cols_raw[3]
        else:
            col_fab = cols_map.get("fabricante") or cols_map.get("marca") or cols_map.get("manufacturer")

        if col_codigo and col_desc:
            for _, row in df.iterrows():
                try:
                    _put(m, row[col_codigo], row[col_desc], row[col_fab] if col_fab else None)
                except Exception:
                    continue
    _articulos_map = m
    return _articulos_map

# -----------------------------
# Permisos por acción (además de roles)
# -----------------------------
ACTIONS = [
    ("dashboard_ver", "Ver dashboard"),
    ("envio_crear", "Crear parte"),
    ("envio_borrar", "Borrar parte"),
    ("instrumento_crear", "Añadir instrumento"),
    ("instrumento_editar", "Editar instrumento"),
    ("instrumento_borrar", "Borrar instrumento"),
    ("fotos_gestionar", "Subir/Borrar fotos"),
    ("estado_cambiar", "Cambiar estado instrumento"),
    ("excel_importar", "Importar Excel"),
    ("usuarios_gestionar", "Gestionar usuarios/roles"),
]

def _user_role(user):
    """Devuelve el role de user soportando dict/sqlite3.Row/objeto."""
    if user is None:
        return None
    # dict-like
    try:
        if hasattr(user, "keys"):
            return user.get("role") if hasattr(user, "get") else user["role"]
    except Exception:
        pass
    # objeto
    try:
        return getattr(user, "role", None)
    except Exception:
        return None

def _user_id(user):
    """Devuelve el id de user soportando dict/sqlite3.Row/objeto."""
    if user is None:
        return None
    try:
        if hasattr(user, "keys"):
            return int(user.get("id") if hasattr(user, "get") else user["id"])
    except Exception:
        pass
    try:
        return int(getattr(user, "id", None))
    except Exception:
        return None

def _default_allowed_by_role(role: str, action: str) -> int:
    role = (role or "").strip().lower()
    if role == "admin":
        return 1
    if role == "recepcion":
        return 1 if action in {"dashboard_ver","envio_crear","instrumento_crear","fotos_gestionar","excel_importar"} else 0
    if role == "tecnico":
        return 1 if action in {"dashboard_ver","instrumento_editar","fotos_gestionar","estado_cambiar"} else 0
    if role == "grabado":
        return 1 if action in {"dashboard_ver"} else 0
    if role == "cliente":
        return 1 if action in {"dashboard_ver"} else 0
    return 0

def _get_user_permissions_map(cur, user_id: int) -> dict:
    cur.execute("SELECT action, allowed FROM user_permissions WHERE user_id=?", (user_id,))
    return {r["action"]: int(r["allowed"] or 0) for r in cur.fetchall()}

def can_action(user, action: str, cur=None) -> bool:
    # Admin siempre
    if _user_role(user) == "admin":
        return True
    role = _user_role(user) or ""
    if cur is None or user is None:
        return _default_allowed_by_role(role, action) == 1
    try:
        cur.execute("SELECT allowed FROM user_permissions WHERE user_id=? AND action=?", (int(_user_id(user)), action))
        row = cur.fetchone()
        if row is None:
            return _default_allowed_by_role(role, action) == 1
        return int(row["allowed"] or 0) == 1
    except Exception:
        # Fallback to default if table missing or other DB error
        return _default_allowed_by_role(role, action) == 1


@app.on_event("startup")
def on_startup():
    # Pre-carga del catálogo de Articulos para evitar LAG en el primer uso
    try:
        from app import load_articulos_map
        load_articulos_map()
        print(">>> [INIT] Catálogo de Artículos pre-cargado con éxito.", flush=True)
    except Exception as e:
        print(f">>> [INIT] Error pre-cargando catálogo: {e}", flush=True)

    init_db()

    # Tabla de permisos por acción (granularidad extra a roles)
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("""
    CREATE TABLE IF NOT EXISTS user_permissions (
      user_id INTEGER NOT NULL,
      action TEXT NOT NULL,
      allowed INTEGER NOT NULL DEFAULT 0,
      PRIMARY KEY (user_id, action),
      FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE
    );
    """)
    conn.commit()
    conn.close()
    # -----------------------------
    # AUTO-RECOVERY ADMIN (solo para demo)
    # Si NO hay ningún admin activo, fuerza:
    #   usuario: admin
    #   password: admin123
    # -----------------------------
    conn = get_conn()
    cur = conn.cursor()

    # Detecta columnas reales en users
    from db import get_table_columns
    cols = get_table_columns(cur, "users")
    colset = set(cols)

    pw_col = "password_hash" if "password_hash" in colset else ("password" if "password" in colset else None)
    has_is_active = "is_active" in colset
    has_created_at = "created_at" in colset
    has_created_at_at = "created_at_at" in colset

    if pw_col:
        if has_is_active:
            cur.execute("SELECT COUNT(*) AS n FROM users WHERE role='admin' AND is_active=1")
        else:
            cur.execute("SELECT COUNT(*) AS n FROM users WHERE role='admin'")
        n_admin = int(cur.fetchone()["n"] or 0)

        if n_admin == 0:
            # No hay admin activo -> forzar admin/admin123
            admin_hash = hash_password("admin123")

            cur.execute("SELECT id FROM users WHERE username='admin'")
            row = cur.fetchone()
            if row:
                # Update existente
                if has_is_active:
                    cur.execute(f"UPDATE users SET role='admin', is_active=1, {pw_col}=? WHERE username='admin'", (admin_hash,))
                else:
                    cur.execute(f"UPDATE users SET role='admin', {pw_col}=? WHERE username='admin'", (admin_hash,))
            else:
                # Insert nuevo compatible con columnas
                cols_ins = ["username", pw_col, "role"]
                vals = ["admin", admin_hash, "admin"]

                if has_is_active:
                    cols_ins.append("is_active")
                    vals.append(1)

                now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                if has_created_at:
                    cols_ins.append("created_at")
                    vals.append(now)
                elif has_created_at_at:
                    cols_ins.append("created_at_at")
                    vals.append(now)

                sql = f"INSERT INTO users ({', '.join(cols_ins)}) VALUES ({', '.join(['?']*len(cols_ins))})"
                cur.execute(sql, tuple(vals))

            conn.commit()

    conn.close()


@app.get("/init_db")
def manual_init_db():
    try:
        on_startup()
        return {"ok": True, "message": "Base de datos inicializada y admin creado (admin/admin123)"}
    except Exception as e:
        return {"ok": False, "error": str(e)}







@app.get('/articulos_lookup')
def articulos_lookup(codigo: str, user=Depends(get_current_user)):
    """Lookup de artículo para autorrelleno en alta manual de instrumentos.

    Devuelve JSON con:
      - found: bool
      - codigo: str (normalizado)
      - descripcion: str
      - fabricante: str (opcional)
    """
    key = _norm_codigo(codigo)
    if not key:
        return {'found': False}

    m = load_articulos_map()
    if not m:
        return {'found': False, 'error': 'Catálogo Articulos no disponible'}

    for k in _codigo_variants(key):
        hit = m.get(k)
        if hit:
            resp = {
                'found': True,
                'codigo': k,
                'descripcion': (hit.get('descripcion') or '').strip(),
            }
            fab = (hit.get('fabricante') or '').strip()
            if fab:
                resp['fabricante'] = fab
            return resp

    return {'found': False}


@app.exception_handler(Exception)
async def global_exception_handler(request: Request, exc: Exception):
    import traceback
    return HTMLResponse(
        f"<h1>Error Global del Servidor</h1><pre>{traceback.format_exc()}</pre>",
        status_code=500
    )

@app.exception_handler(HTTPException)
async def http_exception_handler(request: Request, exc: HTTPException):
    if exc.status_code == 401:
        accept = request.headers.get("accept", "")
        if "text/html" in accept:
            return RedirectResponse(url="/login", status_code=303)
    return JSONResponse({"detail": exc.detail}, status_code=exc.status_code)


def _try_delete_public_photo(public_path: str | None) -> None:
    if not public_path:
        return
    if not public_path.startswith("/static/fotos/"):
        return
    filename = public_path.replace("/static/fotos/", "")
    path_fs = os.path.join(FOTOS_DIR, filename)
    try:
        if os.path.exists(path_fs):
            os.remove(path_fs)
    except Exception:
        pass


def _next_ot_num(cur) -> str:
    yy = datetime.now().year % 100
    prefix = f"{yy:02d}OT"
    cur.execute(
        "SELECT ot_num FROM envios WHERE ot_num LIKE ? ORDER BY ot_num DESC LIMIT 1;",
        (prefix + "%",),
    )
    row = cur.fetchone()
    max_seq = 0
    if row and row["ot_num"]:
        try:
            max_seq = int(row["ot_num"][-5:])
        except Exception:
            max_seq = 0
    return f"{prefix}{(max_seq + 1):05d}"


# -----------------------------
# AUTH
# -----------------------------
@app.get("/login", response_class=HTMLResponse)
def login_form(request: Request):
    return templates.TemplateResponse("login.html", {"request": request, "error": None})


from mail_utils import send_finish_notification, send_credentials_request

@app.post("/solicitar_acceso")
async def handle_solicitud_acceso(request: Request):
    form = await request.form()
    nombre = (form.get("nombre") or "").strip()
    centro = (form.get("centro") or "").strip()
    telefono = (form.get("telefono") or "").strip()
    email = (form.get("email") or "").strip()

    if not all([nombre, centro, email]):
        return RedirectResponse(url="/login?err=missing_fields", status_code=303)

    success, msg = send_credentials_request(nombre, centro, telefono, email)
    if success:
        return RedirectResponse(url="/login?msg=solicitud_enviada", status_code=303)
    else:
        return RedirectResponse(url=f"/login?err={msg}", status_code=303)

@app.post("/login")
async def login(request: Request):
    form = await request.form()
    username = (form.get("username") or "").strip()
    password = form.get("password") or ""

    conn = get_conn()
    cur = conn.cursor()

    schema = _users_schema(cur)
    pw_col = schema.get("password_col")
    if not pw_col:
        conn.close()
        return templates.TemplateResponse("login.html", {"request": request, "error": "Credenciales inválidas"})

    if schema.get("has_is_active"):
        cur.execute(f"SELECT id, {pw_col} AS pw, is_active FROM users WHERE username=?", (username,))
    else:
        cur.execute(f"SELECT id, {pw_col} AS pw, 1 AS is_active FROM users WHERE username=?", (username,))

    u = cur.fetchone()
    conn.close()

    if not u:
        return templates.TemplateResponse("login.html", {"request": request, "error": "Credenciales inválidas"})

    if int(u["is_active"] or 0) == 0:
        return templates.TemplateResponse("login.html", {"request": request, "error": "Usuario desactivado (contacta con el administrador)"})

    if not verify_password(password, u["pw"]):
        return templates.TemplateResponse("login.html", {"request": request, "error": "Credenciales inválidas"})

    response = RedirectResponse(url="/", status_code=303)
    token = sign_session(app.state.serializer, user_id=int(u["id"]))
    response.set_cookie("xurgical_session", token, httponly=True, samesite="lax")
    return response




@app.post("/logout")
def logout():
    resp = RedirectResponse(url="/login", status_code=303)
    resp.delete_cookie("xurgical_session")
    return resp


@app.get("/health")
@app.head("/")
def health_check():
    return {"status": "ok"}

# -----------------------------
# DASHBOARD
# -----------------------------
@app.get("/", response_class=HTMLResponse)
def dashboard(request: Request, user=Depends(get_current_user)):
    # Comprobamos el entorno en tiempo real para el indicador
    is_pg = bool(os.environ.get("DATABASE_URL"))
    is_disk = bool(os.environ.get("XURGICAL_DB_PATH"))
    
    if is_pg:
        db_type = "PRO PG" 
    elif is_disk:
        db_type = "PRO DISK"
    else:
        db_type = "TEMP"
    conn = get_conn()
    cur = conn.cursor()

    q = (request.query_params.get('q') or '').strip()

    # --- 0. Notificaciones de Recogida ---
    n_peticiones_pendientes = 0
    peticiones_recogida = []
    if _user_role(user) in ["admin", "recepcion"]:
        cur.execute("""
            SELECT pr.*, c.nombre as cliente_nombre 
            FROM peticiones_recogida pr
            JOIN clientes c ON pr.cliente_id = c.id
            WHERE pr.estado = 'Pendiente'
            ORDER BY pr.creado_en DESC
        """)
        peticiones_recogida = [dict(r) for r in cur.fetchall()]
        n_peticiones_pendientes = len(peticiones_recogida)

    # --- 0b. Consultas Técnicas (NUEVO Chat) ---
    consultas_list = []
    n_consultas_pendientes = 0
    n_consultas_activas = 0
    if _user_role(user) in ["admin", "recepcion", "tecnico"]:
        cur.execute("""
            SELECT c.*, cl.nombre as cliente_nombre 
            FROM consultas c
            JOIN clientes cl ON c.cliente_id = cl.id
            WHERE c.estado != 'Cerrada'
            ORDER BY c.actualizado_en DESC
        """)
        consultas_list = [dict(r) for r in cur.fetchall()]
        # El staff ve como "pendientes" las que están en estado 'Abierta' (esperando su respuesta)
        n_consultas_pendientes = sum(1 for c in consultas_list if c["estado"] == "Abierta")
        # Pero añadimos una variable para saber cuántas hay en total sin cerrar
        n_consultas_activas = len(consultas_list)
    elif _user_role(user) == "cliente" and user.get("cliente_id"):
        cur.execute("""
            SELECT * FROM consultas 
            WHERE cliente_id = ? AND estado != 'Cerrada'
            ORDER BY actualizado_en DESC
        """, (int(user.get("cliente_id") or 0),))
        consultas_list = [dict(r) for r in cur.fetchall()]
        # El cliente ve como "pendientes" las que han sido 'Respondida' por el SAT
        n_consultas_pendientes = sum(1 for c in consultas_list if c["estado"] == "Respondida")
        n_consultas_activas = len(consultas_list)

    if _user_role(user) in ["admin", "recepcion"]:
        cur.execute("DELETE FROM instrumentos WHERE envio_id IS NULL OR envio_id = 0 OR envio_id NOT IN (SELECT id FROM envios)")
        conn.commit()

    # --- 1. KPIs Globales (Instrumentos) ---
    kpi_where = ""
    kpi_params = []
    if _user_role(user) == "cliente" and user.get("cliente_id"):
        kpi_where = "WHERE e.cliente_id = ?"
        kpi_params = [int(user["cliente_id"])]

    cur.execute(f"""
        SELECT 
            COUNT(i.id) as total,
            SUM(CASE WHEN i.estado='Pendiente' THEN 1 ELSE 0 END) as pendientes,
            SUM(CASE WHEN i.estado='En proceso' THEN 1 ELSE 0 END) as en_proceso,
            SUM(CASE WHEN i.estado='Reparado' THEN 1 ELSE 0 END) as reparados,
            SUM(CASE WHEN i.estado='Baja' THEN 1 ELSE 0 END) as baja
        FROM instrumentos i
        JOIN envios e ON e.id = i.envio_id
        {kpi_where}
    """, tuple(kpi_params))
    
    row_kpi = cur.fetchone()
    if row_kpi:
        kpis = {
            "total": row_kpi["total"] or 0,
            "pendientes": row_kpi["pendientes"] or 0,
            "en_proceso": row_kpi["en_proceso"] or 0,
            "reparado": row_kpi["reparados"] or 0,
            "baja": row_kpi["baja"] or 0,
        }
    else:
        kpis = {"total":0, "pendientes":0, "en_proceso":0, "reparado":0, "baja":0}

    # --- 2. KPIs Partes (Abiertos/Cerrados) ---
    cur.execute(f"SELECT COUNT(*) AS n FROM envios e {kpi_where}", tuple(kpi_params))
    total_partes = cur.fetchone()["n"]

    has_tipo = _envios_has_column(cur, "tipo_trabajo")
    tipo_col = "e.tipo_trabajo" if has_tipo else "'REPARACION'"
    
    cur.execute(f"""
        SELECT 
            e.id,
            {tipo_col} as tipo,
            COUNT(i.id) as total_inst,
            SUM(CASE WHEN COALESCE(i.grabado,0)=1 THEN 1 ELSE 0 END) as n_grabados,
            SUM(CASE WHEN i.estado IN ('Reparado','Baja') THEN 1 ELSE 0 END) as n_terminados
        FROM envios e
        LEFT JOIN instrumentos i ON i.envio_id = e.id
        {kpi_where}
        GROUP BY e.id, {tipo_col}
    """, tuple(kpi_params))
    
    abiertos = 0
    cerrados = 0
    
    rows = cur.fetchall() 
    for r in rows:
        t_inst = r["total_inst"]
        if t_inst == 0:
            abiertos += 1 
            continue
            
        tipo = (r["tipo"] or "REPARACION").upper()
        if tipo == "TRAZABILIDAD":
            if r["n_grabados"] == t_inst:
                cerrados += 1
            else:
                abiertos += 1
        else:
            if r["n_terminados"] == t_inst:
                cerrados += 1
            else:
                abiertos += 1

    kpis_partes = {"total": total_partes, "abiertos": abiertos, "cerrados": cerrados}

    # ✅ AÑADIDO: n_fotos_completas y n_con_alguna_foto para calcular el punto por parte
    # Compatibilidad: en BDs antiguas puede no existir e.tipo_trabajo. En ese caso asumimos REPARACION.
    select_tipo = "e.tipo_trabajo" if _envios_has_column(cur, "tipo_trabajo") else "'REPARACION' AS tipo_trabajo"

    # --- Buscador (OT / Cliente / DataMatrix) ---
    where_clauses = []
    params_q: list = []

    # Filtro por rol 'cliente'
    if _user_role(user) == "cliente" and user.get("cliente_id"):
        where_clauses.append("e.cliente_id = ?")
        params_q.append(int(user["cliente_id"]))
    elif _user_role(user) == "cliente":
        # Si es cliente pero no tiene ID asignado, por seguridad no ve nada
        where_clauses.append("1=0")

    if q:
        is_pg = os.environ.get("DATABASE_URL") is not None
        if is_pg:
            where_clauses.append("(e.ot_num ILIKE ? OR e.cliente ILIKE ? OR i.codigo_datamatrix ILIKE ? OR i.num_serie ILIKE ? OR i.codigo_producto ILIKE ? OR EXISTS (SELECT 1 FROM peticiones_recogida pr WHERE pr.num_peticion ILIKE ? AND pr.cliente_id=e.cliente_id))")
        else:
            # SQLite case-insensitive search (default for LIKE on ASCII)
            where_clauses.append("(e.ot_num LIKE ? OR e.cliente LIKE ? OR i.codigo_datamatrix LIKE ? OR i.num_serie LIKE ? OR i.codigo_producto LIKE ? OR EXISTS (SELECT 1 FROM peticiones_recogida pr WHERE pr.num_peticion LIKE ? AND pr.cliente_id=e.cliente_id))")
        params_q.extend([f"%{q}%", f"%{q}%", f"%{q}%", f"%{q}%", f"%{q}%", f"%{q}%"])

    where_q = "WHERE " + " AND ".join(where_clauses) if where_clauses else ""

    cur.execute(f"""
        SELECT 
            e.*,
            COUNT(i.id) AS n_instrumentos,
            SUM(CASE WHEN 
                (i.foto_entrada_1 IS NOT NULL AND i.foto_entrada_1 != '') OR 
                (i.foto_entrada_2 IS NOT NULL AND i.foto_entrada_2 != '') OR 
                (i.foto_entrada_3 IS NOT NULL AND i.foto_entrada_3 != '') OR 
                (i.foto_entrada_4 IS NOT NULL AND i.foto_entrada_4 != '') OR 
                (i.foto_entrada_5 IS NOT NULL AND i.foto_entrada_5 != '') OR 
                (i.foto_entrada_6 IS NOT NULL AND i.foto_entrada_6 != '') THEN 1 ELSE 0 END) AS n_con_alguna_foto,
            SUM(CASE WHEN COALESCE(i.grabado,0)=1 THEN 1 ELSE 0 END) AS n_grabados,
            SUM(CASE WHEN i.estado = 'En proceso' THEN 1 ELSE 0 END) AS n_en_proceso,
            SUM(CASE WHEN COALESCE(i.repuesto_precio, 0) > 0 THEN 1 ELSE 0 END) AS n_con_repuesto,
            -- Cálculo de 'done' según tipo de trabajo
            SUM(CASE 
                WHEN COALESCE(e.tipo_trabajo,'REPARACION') = 'TRAZABILIDAD' THEN (CASE WHEN COALESCE(i.grabado,0)=1 THEN 1 ELSE 0 END)
                ELSE (CASE WHEN COALESCE(i.estado,'') IN ('Reparado','Baja') THEN 1 ELSE 0 END)
            END) AS n_done
        FROM envios e
        LEFT JOIN instrumentos i ON i.envio_id = e.id
        {where_q}
        GROUP BY e.id
        ORDER BY e.id DESC
        LIMIT 200
    """, tuple(params_q))

    envios = []
    for r in cur.fetchall():
        d = dict(r)
        
        total = int(d.get("n_instrumentos") or 0)
        done = int(d.get("n_done") or 0)
        
        # Lógica de cierre y pendientes
        d["is_closed"] = (total > 0 and done == total)
        d["n_pendientes"] = max(total - done, 0)
        d["color"] = "green" if d["is_closed"] else "red"

        # Indicador de fotos (dot)
        n_con_alguna_foto = int(d.get("n_con_alguna_foto") or 0)
        if total == 0 or n_con_alguna_foto == 0:
            d["foto_dot"] = "red"
        elif n_con_alguna_foto == total:
            d["foto_dot"] = "green"
        else:
            d["foto_dot"] = "yellow"

        envios.append(d)

    # Ordenar: primero abiertas (rojo), luego cerradas. Dentro de cada grupo, ID descendente.
    envios.sort(key=lambda x: (1 if x.get('is_closed') else 0, -int(x.get('id') or 0)))

    # --- Usuarios modal y datos auxiliares ---
    open_users_modal = (request.query_params.get("users") == "1")
    users_list = []
    perms_by_user = {}
    clientes_list_global = []

    if _user_role(user) == "admin":
        schema = _users_schema(cur)
        cur.execute(_select_users_sql(schema))
        users_list = [dict(r) for r in cur.fetchall()]
        for u in users_list:
            perms_by_user[int(u["id"])] = _get_user_permissions_map(cur, int(u["id"]))
        clientes_list_global = _list_clientes(cur)

    # --- Búsqueda de RECOGIDAS ---
    found_recogidas = []
    if q:
        rec_where = ["(num_peticion LIKE ? OR observaciones LIKE ? OR contacto LIKE ?)"]
        rec_params = [f"%{q}%", f"%{q}%", f"%{q}%"]
        if _user_role(user) == "cliente":
            rec_where.append("cliente_id = ?")
            rec_params.append(int(user.get("cliente_id") or 0))
        
        cur.execute(f"""
            SELECT pr.*, c.nombre as cliente_nombre 
            FROM peticiones_recogida pr
            JOIN clientes c ON pr.cliente_id = c.id
            WHERE {" AND ".join(rec_where)}
            ORDER BY pr.creado_en DESC
        """, tuple(rec_params))
        found_recogidas = [dict(r) for r in cur.fetchall()]

    conn.close()
    context = {
        "request": request,
        "user": user,
        "kpis": kpis,
        "kpis_partes": kpis_partes,
        "envios": envios,
        "q": q,
        "open_users_modal": open_users_modal,
        "users_list": users_list,
        "actions": ACTIONS,
        "perms_by_user": perms_by_user,
        "db_type": db_type,
        "clientes_list_global": clientes_list_global,
        "n_peticiones_pendientes": n_peticiones_pendientes,
        "peticiones_recogida": peticiones_recogida,
        "found_recogidas": found_recogidas,
        "consultas_list": consultas_list,
        "n_consultas_pendientes": n_consultas_pendientes,
        "n_consultas_activas": n_consultas_activas or 0,
    }

    return templates.TemplateResponse(
        "dashboard.html",
        context,
    )


# -----------------------------
# TRAZABILIDAD / HISTORIAL
# -----------------------------
@app.get("/api/trazabilidad/buscar")
def api_trazabilidad_buscar(q: str = "", user=Depends(get_current_user)):
    """Busca el historial de un instrumento por DataMatrix o Nº de Serie."""
    q = (q or "").strip()
    if not q:
        return {"results": []}

    conn = get_conn()
    cur = conn.cursor()

    where_clauses = ["(i.codigo_datamatrix = ? OR i.num_serie = ? OR i.nombre_trazabilidad = ?)"]
    params = [q, q, q]

    # Filtro por cliente
    if _user_role(user) == "cliente":
        u_cli_id = (user.get("cliente_id") if isinstance(user, dict) else getattr(user, "cliente_id", None))
        if u_cli_id:
            where_clauses.append("e.cliente_id = ?")
            params.append(int(u_cli_id))
        else:
            conn.close()
            return {"results": []}

    where_sql = " AND ".join(where_clauses)
    
    sql = f"""
        SELECT i.id, i.envio_id, i.denominacion, i.codigo_producto, i.fabricante, i.estado, i.creado_en,
               e.ot_num, e.fecha, e.tipo_trabajo
        FROM instrumentos i
        JOIN envios e ON e.id = i.envio_id
        WHERE {where_sql}
        ORDER BY i.creado_en DESC
        LIMIT 50
    """
    
    try:
        cur.execute(sql, tuple(params))
        rows = cur.fetchall()
        results = [dict(r) for r in rows]
        
        # Formatear fechas para el JSON
        for r in results:
            if r.get("fecha"):
                r["fecha_fmt"] = format_fecha(r["fecha"])
            if r.get("creado_en"):
                r["creado_fmt"] = format_fecha(r["creado_en"])
                
        return {"results": results}
    except Exception as e:
        print(f"Error en api_trazabilidad_buscar: {e}")
        return {"results": [], "error": str(e)}
    finally:
        conn.close()


@app.post("/perfil/password")
def change_own_password(
    password: str = Form(...),
    user=Depends(get_current_user),
):
    if not (password or "").strip() or len(password) < 6:
        return RedirectResponse(url="/?err=pw_too_short", status_code=303)

    conn = get_conn()
    cur = conn.cursor()
    schema = _users_schema(cur)
    pw_col = schema.get("password_col")
    
    user_id = (user.get("id") if isinstance(user, dict) else getattr(user, "id", None))
    
    if not pw_col or not user_id:
        conn.close()
        return RedirectResponse(url="/?err=db", status_code=303)

    cur.execute(f"UPDATE users SET {pw_col}=? WHERE id=?", (hash_password(password), user_id))
    conn.commit()
    conn.close()
    
    # Redirigir a login para que vuelva a entrar con la nueva clave
    response = RedirectResponse(url="/login?msg=pw_changed", status_code=303)
    response.delete_cookie("access_token")
    return response


# -----------------------------
# EXPORTACIÓN
# -----------------------------
def _bool_param(v: str | None) -> bool:
    if v is None:
        return False
    return str(v).strip().lower() in {"1", "true", "t", "yes", "y", "on"}


def _build_instrumentos_where(estado: str | None, solo_pendientes: bool, solo_grabados: bool | None):
    where = []
    params: list = []

    if solo_pendientes:
        where.append("i.estado IN ('Pendiente','En proceso')")
    elif estado and estado != "TODOS":
        where.append("i.estado = ?")
        params.append(estado)

    if solo_grabados is True:
        where.append("COALESCE(i.grabado,0)=1")
    elif solo_grabados is False:
        where.append("COALESCE(i.grabado,0)=0")

    sql = " AND ".join(where) if where else "1=1"
    return sql, params


@app.get("/export", response_class=HTMLResponse)
def export_home(request: Request, user=Depends(require_roles("admin", "recepcion", "cliente"))):
    # Página con opciones de exportación
    conn = get_conn()
    cur = conn.cursor()
    
    if _user_role(user) == "cliente" and user.get("cliente_id"):
        cur.execute("SELECT id, ot_num, cliente, fecha FROM envios WHERE cliente_id=? ORDER BY id DESC LIMIT 400", (int(user["cliente_id"]),))
    elif _user_role(user) == "cliente":
        # Por seguridad
        cur.execute("SELECT id, ot_num, cliente, fecha FROM envios WHERE 1=0")
    else:
        cur.execute("SELECT id, ot_num, cliente, fecha FROM envios ORDER BY id DESC LIMIT 400")
        
    envios = [dict(r) for r in cur.fetchall()]
    conn.close()
    return templates.TemplateResponse(
        "export.html",
        {
            "request": request,
            "user": user,
            "envios": envios,
            "estados": ["TODOS", "Pendiente", "En proceso", "Reparado", "Baja"],
        },
    )


@app.get("/guias", response_class=HTMLResponse)
def guias_cliente(request: Request, user=Depends(get_current_user)):
    return templates.TemplateResponse("guias.html", {"request": request, "user": user})


@app.get("/export/download")
def export_download(
    request: Request,
    user=Depends(require_roles("admin", "recepcion", "cliente")),
    scope: str = "partes",  # partes | instrumentos | parte
    envio_id: str | None = None,
    estado: str | None = None,
    solo_pendientes: str | None = None,
    grabado: str | None = None,  # todos|si|no
    fmt: str = "xlsx",  # xlsx|csv
):
    fmt = (fmt or "xlsx").lower()
    if fmt not in {"xlsx", "csv"}:
        raise HTTPException(status_code=400, detail="Formato no soportado")

    scope = (scope or "partes").lower()
    if scope not in {"partes", "instrumentos", "parte"}:
        raise HTTPException(status_code=400, detail="Scope no soportado")

    solo_pend = _bool_param(solo_pendientes)
    
    # Parse envio_id a int si viene
    eid: int | None = None
    if envio_id and str(envio_id).strip():
        try:
            eid = int(envio_id)
        except:
            raise HTTPException(status_code=400, detail="envio_id debe ser un número")

    grabado = (grabado or "todos").lower()
    solo_grabados: bool | None = None
    if grabado == "si":
        solo_grabados = True
    elif grabado == "no":
        solo_grabados = False

    cli_id = (user.get("cliente_id") if isinstance(user, dict) else getattr(user, "cliente_id", None))
    is_cliente = _user_role(user) == "cliente"

    conn = get_conn()
    cur = conn.cursor()

    # ---- DATOS ----
    partes_rows: list[dict] = []
    inst_rows: list[dict] = []

    if scope in {"partes", "parte"}:
        if scope == "parte" and not eid:
            raise HTTPException(status_code=400, detail="Falta envio_id")

        where_env = "1=1"
        params_env: list = []
        if eid:
            where_env = "e.id=?"
            params_env = [eid]
            
        if is_cliente:
            if cli_id:
                where_env += " AND e.cliente_id=?"
                params_env.append(int(cli_id))
            else:
                where_env += " AND 1=0"

        cur.execute(
            f"""
            SELECT
                e.id,
                e.ot_num,
                e.cliente,
                e.fecha,
                e.creado_en,
                COUNT(i.id) AS n_instrumentos,
                SUM(CASE WHEN i.estado IN ('Pendiente','En proceso') THEN 1 ELSE 0 END) AS n_pendientes,
                SUM(CASE WHEN i.estado='Reparado' THEN 1 ELSE 0 END) AS n_reparados,
                SUM(CASE WHEN i.estado='Baja' THEN 1 ELSE 0 END) AS n_baja,
                SUM(CASE WHEN COALESCE(i.grabado,0)=1 THEN 1 ELSE 0 END) AS n_grabados,
                SUM(CASE WHEN i.foto_entrada_1 IS NOT NULL AND i.foto_entrada_2 IS NOT NULL AND i.foto_entrada_3 IS NOT NULL AND i.foto_entrada_4 IS NOT NULL AND i.foto_entrada_5 IS NOT NULL AND i.foto_entrada_6 IS NOT NULL THEN 1 ELSE 0 END) AS n_fotos_completas
            FROM envios e
            LEFT JOIN instrumentos i ON i.envio_id = e.id
            WHERE {where_env}
            GROUP BY e.id
            ORDER BY e.id DESC
            """,
            tuple(params_env),
        )
        for r in cur.fetchall():
            d = dict(r)
            n_inst = int(d.get("n_instrumentos") or 0)
            n_pend = int(d.get("n_pendientes") or 0)
            d["cerrado"] = 1 if (n_inst > 0 and n_pend == 0) else 0
            partes_rows.append(d)

    if scope in {"instrumentos", "parte"}:
        where_i, params_i = _build_instrumentos_where(estado, solo_pend, solo_grabados)
        where_envio = "1=1"
        params_envio: list = []
        if scope == "parte":
            where_envio = "i.envio_id=?"
            params_envio = [eid] # Usamos eid ya parseado
            
        if is_cliente:
            if cli_id:
                where_envio += " AND e.cliente_id=?"
                params_envio.append(int(cli_id))
            else:
                where_envio += " AND 1=0"

        cur.execute(
            f"""
            SELECT
                i.id,
                i.envio_id,
                e.ot_num,
                e.cliente,
                e.fecha,
                i.codigo_datamatrix,
                i.fabricante,
                i.codigo_producto,
                i.denominacion,
                i.num_serie,
                i.estado,
                COALESCE(i.grabado,0) AS grabado,
                i.grabado_en,
                i.grabado_por,
                i.foto_entrada_1,
                i.foto_entrada_2,
                i.foto_entrada_3,
                i.foto_entrada_4,
                i.foto_entrada_5,
                i.foto_entrada_6,
                i.observaciones,
                i.creado_en,
                i.actualizado_en
            FROM instrumentos i
            JOIN envios e ON e.id = i.envio_id
            WHERE {where_envio} AND {where_i}
            ORDER BY e.id DESC, i.id ASC
            """,
            tuple(params_envio + params_i),
        )
        inst_rows = [dict(r) for r in cur.fetchall()]

    conn.close()

    # ---- RESPUESTA (CSV/XLSX) ----
    now_tag = datetime.now().strftime("%Y%m%d_%H%M")
    base_name = f"xurgical_export_{scope}_{now_tag}"

    if fmt == "csv":
        # Para CSV exportamos una sola tabla (prioridad instrumentos si aplica)
        rows = inst_rows if scope in {"instrumentos", "parte"} else partes_rows
        if not rows:
            rows = []
        headers = list(rows[0].keys()) if rows else []
        sio = io.StringIO()
        w = csv.writer(sio, delimiter=";")
        w.writerow(headers)
        for r in rows:
            w.writerow([r.get(h, "") for h in headers])

        data = sio.getvalue().encode("utf-8-sig")
        return StreamingResponse(
            io.BytesIO(data),
            media_type="text/csv; charset=utf-8",
            headers={"Content-Disposition": f"attachment; filename={base_name}.csv"},
        )

    # XLSX
    if Workbook is None:
        raise HTTPException(status_code=500, detail="openpyxl no está disponible")

    wb = Workbook()
    # elimina la hoja por defecto
    ws0 = wb.active
    wb.remove(ws0)

    def add_sheet(title: str, rows: list[dict]):
        ws = wb.create_sheet(title=title)
        headers = list(rows[0].keys()) if rows else []
        ws.append(headers)
        for r in rows:
            ws.append([r.get(h, "") for h in headers])
        # ancho columnas
        for col_idx, h in enumerate(headers, start=1):
            maxlen = len(str(h))
            for rr in rows[:200]:
                maxlen = max(maxlen, len(str(rr.get(h, "")) or ""))
            ws.column_dimensions[chr(64 + col_idx)].width = min(max(10, maxlen + 2), 55)

    if partes_rows:
        add_sheet("Partes", partes_rows)
    if inst_rows:
        add_sheet("Instrumentos", inst_rows)

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={base_name}.xlsx"},
    )
# -----------------------------
# ENVÍOS
# -----------------------------

# -----------------------------
# CLIENTES (admin/recepcion)
# -----------------------------
@app.get("/clientes", response_class=HTMLResponse)
def clientes_list(request: Request, user=Depends(require_roles("admin", "recepcion"))):
    conn = get_conn()
    cur = conn.cursor()
    clientes = _list_clientes(cur)
    conn.close()
    return templates.TemplateResponse(
        "clientes_list.html",
        {"request": request, "user": user, "clientes": clientes},
    )


@app.get("/clientes/nuevo", response_class=HTMLResponse)
def clientes_nuevo_form(request: Request, user=Depends(require_roles("admin", "recepcion"))):
    return templates.TemplateResponse(
        "clientes_form.html",
        {"request": request, "user": user, "mode": "new", "cliente": None},
    )


@app.post("/clientes/nuevo")
def clientes_nuevo_crear(
    nombre: str = Form(""),
    prefijo: str = Form(""),
    email: str = Form(""),
    prefijo_nombre: str = Form(""),
    ultimo_numero: int = Form(0),
    user=Depends(require_roles("admin", "recepcion")),
):
    nombre = (nombre or "").strip()
    if not nombre:
        return RedirectResponse(url="/clientes?err=nombre", status_code=303)

    conn = get_conn()
    cur = conn.cursor()

    # Prevenir duplicados (validación manual además del índice UNIQUE)
    cur.execute("SELECT id FROM clientes WHERE nombre = ?", (nombre,))
    if cur.fetchone():
        conn.close()
        return RedirectResponse(url="/clientes?err=exists", status_code=303)

    cur.execute(
        "INSERT INTO clientes (nombre, prefijo, email, prefijo_nombre, ultimo_numero) VALUES (?, ?, ?, ?, ?)",
        (nombre, (prefijo or "").strip(), (email or "").strip(), (prefijo_nombre or "").strip(), int(ultimo_numero or 0)),
    )
    conn.commit()
    conn.close()
    return RedirectResponse(url="/clientes", status_code=303)


@app.get("/clientes/{cliente_id}/editar", response_class=HTMLResponse)
def clientes_editar_form(request: Request, cliente_id: int, user=Depends(require_roles("admin", "recepcion"))):
    conn = get_conn()
    cur = conn.cursor()
    cli = _get_cliente(cur, cliente_id)
    conn.close()
    if not cli:
        return HTMLResponse("Cliente no encontrado", status_code=404)
    return templates.TemplateResponse(
        "clientes_form.html",
        {"request": request, "user": user, "mode": "edit", "cliente": cli},
    )


@app.post("/clientes/{cliente_id}/editar")
def clientes_editar_guardar(
    cliente_id: int,
    nombre: str = Form(""),
    prefijo: str = Form(""),
    email: str = Form(""),
    prefijo_nombre: str = Form(""),
    ultimo_numero: int = Form(0),
    user=Depends(require_roles("admin", "recepcion")),
):
    nombre = (nombre or "").strip()
    if not nombre:
        return RedirectResponse(url=f"/clientes/{cliente_id}/editar?err=nombre", status_code=303)
    conn = get_conn()
    cur = conn.cursor()
    cur.execute(
        "UPDATE clientes SET nombre=?, prefijo=?, email=?, prefijo_nombre=?, ultimo_numero=? WHERE id=?",
        (nombre, (prefijo or "").strip(), (email or "").strip(), (prefijo_nombre or "").strip(), int(ultimo_numero or 0), int(cliente_id)),
    )
    conn.commit()
    conn.close()
    return RedirectResponse(url="/clientes", status_code=303)

@app.get("/envios/nuevo", response_class=HTMLResponse)
def nuevo_envio_form(request: Request, user=Depends(require_roles("admin", "recepcion"))):
    conn = get_conn()
    cur = conn.cursor()
    clientes = _list_clientes(cur)
    conn.close()
    return templates.TemplateResponse(
        "envio_nuevo.html",
        {"request": request, "user": user, "clientes": clientes},
    )


@app.post("/envios/nuevo")
def nuevo_envio_crear(
    referencia: str = Form(""),
    cliente_id: str = Form(""),
    cliente: str = Form(""),
    tipo_trabajo: str = Form("REPARACION"),
    fecha: str = Form(""),
    observaciones: str = Form(""),
    user=Depends(require_roles("admin", "recepcion")),
):
    try:
        # Validación de fecha obligatoria
        fecha = (fecha or "").strip()
        if not fecha:
            return RedirectResponse(url="/envios/nuevo?err=fecha", status_code=303)

        # Permiso granular (además del rol)
        conn_perm = get_conn()
        cur_perm = conn_perm.cursor()
        if not can_action(user, "envio_crear", cur_perm):
            conn_perm.close()
            return RedirectResponse(url="/?err=perm", status_code=303)
        conn_perm.close()

        conn = get_conn()
        cur = conn.cursor()

        ot_num = _next_ot_num(cur)

        tipo_trabajo = (tipo_trabajo or "REPARACION").strip().upper()
        if tipo_trabajo not in ("REPARACION", "TRAZABILIDAD", "OPTICA_RIGIDA"):
            tipo_trabajo = "REPARACION"

        cli_id_val = None
        cli_nombre = (cliente or "").strip()
        if cliente_id:
            try:
                cli_id_val = int(cliente_id)
            except Exception:
                cli_id_val = None

        # Si viene un cliente_id, lo resolvemos a nombre (y garantizamos que existe)
        if cli_id_val:
            cli = _get_cliente(cur, cli_id_val)
            if not cli:
                conn.close()
                return RedirectResponse(url="/envios/nuevo?err=cliente", status_code=303)
            cli_nombre = cli["nombre"]
        else:
            # Si no hay ID, debe haber un nombre manual
            if not cli_nombre:
                conn.close()
                return RedirectResponse(url="/envios/nuevo?err=cliente", status_code=303)
            
            # Si es trazabilidad, requerimos cliente registrado (ID).
            if tipo_trabajo == "TRAZABILIDAD":
                conn.close()
                return RedirectResponse(url="/envios/nuevo?err=cliente", status_code=303)

        sql = """
            INSERT INTO envios (ot_num, nombre_archivo, cliente, cliente_id, tipo_trabajo, fecha, observaciones)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        """
        vals = [
            ot_num,
            (referencia or "").strip(),
            cli_nombre,
            cli_id_val,
            tipo_trabajo,
            (fecha or "").strip(),
            (observaciones or "").strip(),
        ]

        is_pg = bool(os.environ.get("DATABASE_URL"))
        if is_pg:
            sql += " RETURNING id"
            cur.execute(sql, tuple(vals))
            row = cur.fetchone()
            if row:
                envio_id = int(row["id"])
            else:
                # Fallback extremo: buscar por OT exacta que acabamos de meter
                cur.execute("SELECT id FROM envios WHERE ot_num=?", (ot_num,))
                row_f = cur.fetchone()
                if row_f:
                    envio_id = int(row_f["id"])
                else:
                     raise Exception("No se pudo obtener el ID del nuevo envío en Postgres")
        else:
            cur.execute(sql, tuple(vals))
            envio_id = cur.lastrowid

        conn.commit()
        conn.close()
        # Al crear una OT, vamos al detalle para añadir instrumentos.
        return RedirectResponse(url=f"/envios/{envio_id}", status_code=303)

    except Exception as e:
        import traceback
        return HTMLResponse(f"<h1>Error al crear envio</h1><pre>{traceback.format_exc()}</pre>", status_code=500)


@app.get("/envios/{envio_id}/editar", response_class=HTMLResponse)
def envio_editar_form(request: Request, envio_id: int, user=Depends(require_roles("admin", "recepcion"))):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("SELECT * FROM envios WHERE id=?", (envio_id,))
    envio = cur.fetchone()
    clientes = _list_clientes(cur)
    conn.close()

    if not envio:
        return HTMLResponse("Envío no encontrado", status_code=404)

    return templates.TemplateResponse(
        "envio_editar.html",
        {
            "request": request,
            "user": user,
            "envio": dict(envio),
            "clientes": clientes
        },
    )


@app.post("/envios/{envio_id}/editar")
def envio_editar_guardar(
    envio_id: int,
    referencia: str = Form(""),
    cliente_id: str = Form(""),
    cliente: str = Form(""),
    tipo_trabajo: str = Form(None),
    fecha: str = Form(None),
    observaciones: str = Form(""),
    user=Depends(require_roles("admin", "recepcion")),
):
    try:
        conn = get_conn()
        cur = conn.cursor()
        
        # Validar existencia
        cur.execute("SELECT * FROM envios WHERE id=?", (envio_id,))
        envio_old = cur.fetchone()
        if not envio_old:
            conn.close()
            return HTMLResponse("Envío no encontrado", status_code=404)

        # Resolución de cliente
        cli_id_val = None
        cli_nombre = (cliente or "").strip()
        if cliente_id:
            try:
                cli_id_val = int(cliente_id)
            except:
                cli_id_val = None
        
        if cli_id_val:
            cli = _get_cliente(cur, cli_id_val)
            if cli:
                cli_nombre = cli["nombre"]
        
        # Si no hay nombre manual ni ID válido, error
        if not cli_nombre and not cli_id_val:
             conn.close()
             return RedirectResponse(url=f"/envios/{envio_id}/editar?err=cliente", status_code=303)

        sql = """
            UPDATE envios 
            SET nombre_archivo=?, cliente=?, cliente_id=?, tipo_trabajo=?, fecha=?, observaciones=?
            WHERE id=?
        """
        params = [
            (referencia or "").strip(),
            cli_nombre,
            cli_id_val,
            (tipo_trabajo or envio_old["tipo_trabajo"]),
            (fecha or envio_old["fecha"]),
            (observaciones or "").strip(),
            envio_id
        ]
        
        cur.execute(sql, tuple(params))
        conn.commit()
        conn.close()
        
        return RedirectResponse(url="/", status_code=303)

    except Exception as e:
        import traceback
        return HTMLResponse(f"<h1>Error al editar envio</h1><pre>{traceback.format_exc()}</pre>", status_code=500)


@app.post("/envios/{envio_id}/toggle_aceptado")
def toggle_aceptado(envio_id: int, user=Depends(get_current_user)):
    # Solo admin o recepcion pueden marcar como aceptado
    if _user_role(user) not in ["admin", "recepcion"]:
        return JSONResponse({"ok": False, "error": "No tienes permisos"}, status_code=403)

    conn = get_conn()
    cur = conn.cursor()
    cur.execute("SELECT aceptado FROM envios WHERE id=?", (envio_id,))
    row = cur.fetchone()
    if not row:
        conn.close()
        return JSONResponse({"ok": False, "error": "Envío no encontrado"}, status_code=404)
    
    nuevo_estado = 1 if int(row["aceptado"] or 0) == 0 else 0
    cur.execute("UPDATE envios SET aceptado=? WHERE id=?", (nuevo_estado, envio_id))
    conn.commit()
    conn.close()
    return {"ok": True, "aceptado": nuevo_estado}


@app.get("/ot/{ot_num}")
def ver_ot_directa(ot_num: str, user=Depends(get_current_user)):
    """Busca una OT por su numero y redirige al detalle."""
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("SELECT id FROM envios WHERE ot_num=?", (ot_num,))
    e = cur.fetchone()
    conn.close()
    if not e:
        return HTMLResponse("OT no encontrada", status_code=404)
    
    # Si es técnico, le llevamos a la vista móvil directamente? O lo hacemos por rol en /envios/{id}?
    # Mejor centralizar en /envios/{id} o crear una ruta específica.
    return RedirectResponse(url=f"/envios/{e['id']}", status_code=303)

@app.post("/envios/{envio_id}/aviso_finalizacion")
def envio_aviso_finalizacion(envio_id: int, user=Depends(require_roles("admin", "recepcion"))):
    """Envía un aviso por email al cliente si el parte está terminado."""
    conn = get_conn()
    cur = conn.cursor()
    
    # 1. Obtener datos del envío y cliente
    cur.execute("""
        SELECT e.*, c.email as cliente_email, c.nombre as cliente_nombre
        FROM envios e
        LEFT JOIN clientes c ON e.cliente_id = c.id
        WHERE e.id=?
    """, (envio_id,))
    envio = cur.fetchone()
    
    if not envio:
        conn.close()
        return {"ok": False, "error": "Envío no encontrado"}
        
    if not envio["cliente_email"]:
        conn.close()
        return {"ok": False, "error": "El cliente no tiene un email configurado para notificaciones"}

    # 2. Verificar si está terminado (todas las piezas reparadas o baja)
    cur.execute("SELECT COUNT(*) as total FROM instrumentos WHERE envio_id=?", (envio_id,))
    total = cur.fetchone()["total"]
    
    cur.execute("SELECT COUNT(*) as done FROM instrumentos WHERE envio_id=? AND estado IN ('Reparado', 'Baja')", (envio_id,))
    done = cur.fetchone()["done"]
    
    if total == 0 or done < total:
        conn.close()
        return {"ok": False, "error": "El parte aún no está totalmente terminado (faltan piezas por revisar)"}

    # 3. Enviar email
    from mail_utils import send_finish_notification
    success, msg = send_finish_notification(
        envio["cliente_email"], 
        envio["cliente_nombre"] or envio["cliente"], 
        envio["ot_num"], 
        total
    )
    
    if success:
        # Marcar como enviado (necesitaremos esta columna en DB)
        try:
            cur.execute("UPDATE envios SET aviso_enviado=1 WHERE id=?", (envio_id,))
            conn.commit()
        except:
            pass # Si falla no bloqueamos la respuesta
            
    conn.close()
    return {"ok": success, "message": msg}


@app.get("/envios/{envio_id}", response_class=HTMLResponse)
def ver_envio(request: Request, envio_id: int, user=Depends(get_current_user)):
    conn = get_conn()
    cur = conn.cursor()

    cur.execute("SELECT * FROM envios WHERE id=?", (envio_id,))
    envio = cur.fetchone()
    if not envio:
        conn.close()
        return HTMLResponse("Envío no encontrado", status_code=404)

    # SEGURIDAD: Si es rol 'cliente', solo puede ver sus propios envíos y si está ACEPTADO
    if _user_role(user) == "cliente":
        u_cli_id = (user.get("cliente_id") if isinstance(user, dict) else getattr(user, "cliente_id", None))
        e_cli_id = envio.get("cliente_id")
        if not u_cli_id or int(e_cli_id or 0) != int(u_cli_id):
            conn.close()
            return HTMLResponse("Acceso denegado: este envío no pertenece a su centro", status_code=403)
        
        # Nueva restricción: debe estar aceptado por recepción/admin para que el cliente lo vea
        if not bool(envio["aceptado"]):
            conn.close()
            return HTMLResponse("<h1>Acceso restringido</h1><p>Este parte aún no ha sido validado/aceptado. Por favor, contacte con recepción para más información.</p>", status_code=403)

    cur.execute("""
        SELECT i.*, 
               (SELECT 1 FROM instrumento_informes ii WHERE ii.instrumento_id = i.id LIMIT 1) as has_archived,
               (SELECT 1 FROM instrumento_qc_optica qc WHERE qc.instrumento_id = i.id LIMIT 1) as has_qc_data,
               (SELECT 1 FROM instrumento_checklist ic WHERE ic.instrumento_id = i.id LIMIT 1) as has_checklist
        FROM instrumentos i 
        WHERE i.envio_id=? 
        ORDER BY i.id DESC
    """, (envio_id,))
    instrumentos = [dict(r) for r in cur.fetchall()]
    
    # Calcular si está totalmente terminado (todas las piezas reparadas o baja)
    total_inst = len(instrumentos)
    done_inst = sum(1 for i in instrumentos if i["estado"] in ["Reparado", "Baja"])
    is_finished = (total_inst > 0 and done_inst == total_inst)

    for r in instrumentos:
        # Limpieza de trazabilidad
        if r.get("nombre_trazabilidad"):
            r["nombre_trazabilidad"] = _clean_trz(r["nombre_trazabilidad"])
        
        # Bandera para mostrar icono PDF (informe)
        r["has_informe"] = bool(r.get("has_archived") or r.get("has_qc_data"))

    # DEF_TRZ_NOMBRE_OK: prepara nombre_trazabilidad SOLO para pantalla de grabación (copiar/pegar)
    envio_dict = dict(envio)
    prefijo = ""
    for k in ("prefijo_nombre", "prefijo_trazabilidad", "prefijo_traz", "prefijo_etiqueta", "prefijo"):
        v = envio_dict.get(k)
        if v:
            prefijo = str(v).strip()
            break

    for r in instrumentos:
        nt = (r.get("nombre_trazabilidad") or "").strip()
        if nt: nt = nt.strip().strip("'").strip('"')
        if nt and prefijo and re.fullmatch(r"\d{5}", nt):
            nt = f"{prefijo}{nt}"
        if (not nt) and prefijo:
            dm = (r.get("codigo_datamatrix") or "").strip()
            if dm:
                try:
                    nt = _build_nombre_trazabilidad(prefijo, dm)
                except Exception:
                    nt = ""
        r["nombre_trazabilidad"] = _clean_trz(nt)

    conn.close()

    if _user_role(user) == "tecnico":
        return templates.TemplateResponse(
            "tecnico_parte.html",
            {"request": request, "user": user, "envio": dict(envio), "instrumentos": instrumentos, "is_finished": is_finished},
        )

    return templates.TemplateResponse(
        "envio_detalle.html",
        {"request": request, "user": user, "envio": dict(envio), "instrumentos": instrumentos, "is_finished": is_finished},
    )


# -----------------------------
# API CHECKLIST (para modo técnico móvil)
# -----------------------------
@app.get("/api/instrumentos/{instrumento_id}/checklist")
def api_get_checklist(instrumento_id: int, user=Depends(get_current_user)):
    conn = get_conn()
    cur = conn.cursor()
    
    # Obtener tipo de OT
    cur.execute("""
        SELECT e.tipo_trabajo 
        FROM instrumentos i 
        JOIN envios e ON e.id = i.envio_id 
        WHERE i.id=?
    """, (instrumento_id,))
    row = cur.fetchone()
    tipo_ot = (row["tipo_trabajo"] if row else "REPARACION") or "REPARACION"

    # Si es TRAZABILIDAD, permitimos usar el checklist de REPARACION
    if tipo_ot == "TRAZABILIDAD":
        tipo_ot = "REPARACION"
    
    cur.execute("""
        SELECT ci.id AS item_id, ci.nombre, 
               COALESCE(ic.hecho,0) AS hecho
        FROM checklist_items ci
        LEFT JOIN instrumento_checklist ic ON ic.item_id = ci.id AND ic.instrumento_id = ?
        WHERE ci.activo = 1 AND ci.tipo_trabajo = ?
        ORDER BY LOWER(ci.nombre) ASC
    """, (instrumento_id, tipo_ot))
    
    checklist = [dict(r) for r in cur.fetchall()]
    conn.close()
    return {"checklist": checklist}

@app.post("/api/instrumentos/{instrumento_id}/checklist")
async def api_save_checklist(instrumento_id: int, request: Request, user=Depends(require_roles("admin", "tecnico"))):
    data = await request.json()
    items_hechos = data.get("items", []) # lista de IDs de items marcados
    
    conn = get_conn()
    cur = conn.cursor()
    
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    username = user.get("username")
    
    # 1. Desmarcar todos para este instrumento (o marcar como 0)
    cur.execute("DELETE FROM instrumento_checklist WHERE instrumento_id=?", (instrumento_id,))
    
    # 2. Insertar los marcados
    for item_id in items_hechos:
        cur.execute("""
            INSERT INTO instrumento_checklist (instrumento_id, item_id, hecho, hecho_por, hecho_en)
            VALUES (?, ?, 1, ?, ?)
        """, (instrumento_id, item_id, username, now))
    
    conn.commit()
    conn.close()
    return {"ok": True}


# -----------------------------
# ETIQUETA (pegatina) OT
# -----------------------------
def _build_etiqueta_pdf(ot_num: str, cliente: str, fecha: str, n_instrumentos: int, referencia: str = "",
                        fabricante: str = "", modelo: str = "", serie: str = "") -> bytes:
    """Genera una etiqueta PDF con texto + código de barras Code128.

    Contenido del barcode: OT|CLIENTE|FECHA|N
    """
    # Etiqueta térmica 29x62 mm (Brother QL-700).
    # Usamos 62mm de ancho y 29mm de alto (paisaje).
    w, h = 62 * mm, 29 * mm
    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=(w, h))

    # Márgenes y posiciones ajustadas para 29mm de alto
    # x0 es el margen izquierdo. Lo ponemos en 7mm para que no se pegue al borde.
    x0 = 7 * mm
    y_top = h - 2 * mm

    # Título OT
    c.setFont("Helvetica-Bold", 10)
    c.drawString(x0, y_top - 3.5 * mm, f"OT: {ot_num}")

    c.setFont("Helvetica", 7.5)
    
    # Recorta texto si es muy largo (ajustado para ancho de 62mm)
    def truncate(s, limit=22):
        s = (s or "").strip()
        return s[:limit-3] + "…" if len(s) > limit else s

    cli = truncate(cliente, 22)
    ref = truncate(referencia, 22)
    fab = truncate(fabricante, 22)
    mod = truncate(modelo, 22)
    sn = truncate(serie, 22)

    # Distribución compacta (line_h = 3mm)
    curr_y = y_top - 7.5 * mm
    step = 3.2 * mm

    c.drawString(x0, curr_y, f"Cli: {cli}")
    curr_y -= step
    
    if fab or mod:
        c.drawString(x0, curr_y, f"Art: {fab} {mod}")
        curr_y -= step
    else:
        c.drawString(x0, curr_y, f"Ref: {ref}")
        curr_y -= step

    if sn:
        c.drawString(x0, curr_y, f"S/N: {sn}")
        curr_y -= step
    else:
        c.drawString(x0, curr_y, f"Fecha: {fecha}")
        curr_y -= step

    c.drawString(x0, curr_y, f"Inst: {n_instrumentos} | {fecha if sn else ''}")

    # QR Code
    payload = str(ot_num)
    from reportlab.graphics.shapes import Drawing
    from reportlab.graphics import renderPDF
    
    qr_code = qr.QrCodeWidget(payload)
    bounds = qr_code.getBounds()
    qr_w = bounds[2] - bounds[0]
    qr_h = bounds[3] - bounds[1]
    
    # QR de 17mm para 29mm de alto
    size = 17 * mm
    d = Drawing(size, size, transform=[size/qr_w, 0, 0, size/qr_h, 0, 0])
    d.add(qr_code)
    
    # Posicionamos el QR a la derecha con margen de 3mm respecto al final de los 62mm
    renderPDF.draw(d, c, w - size - 3 * mm, 3.5 * mm)
    
    c.showPage()
    c.save()
    return buf.getvalue()


@app.get("/envios/{envio_id}/etiqueta.pdf")
def etiqueta_envio(envio_id: int, user=Depends(get_current_user)):
    """Devuelve una pegatina PDF para la OT."""
    conn = get_conn()
    cur = conn.cursor()
    # Intentamos detectar si existe la columna tipo_trabajo
    cols = []
    try:
        from db import get_table_columns
        cols = get_table_columns(cur, "envios")
    except: pass

    if "tipo_trabajo" in cols:
        cur.execute("SELECT id, ot_num, cliente, fecha, cliente_id, nombre_archivo, tipo_trabajo FROM envios WHERE id=?", (envio_id,))
    else:
        cur.execute("SELECT id, ot_num, cliente, fecha, cliente_id, nombre_archivo FROM envios WHERE id=?", (envio_id,))
    
    e = cur.fetchone()
    if not e:
        conn.close()
        return HTMLResponse("OT no encontrada", status_code=404)

    # Seguridad para clientes
    if _user_role(user) == "cliente":
        u_cli_id = (user.get("cliente_id") if isinstance(user, dict) else getattr(user, "cliente_id", None))
        if not u_cli_id or int(e["cliente_id"] or 0) != int(u_cli_id):
            conn.close()
            return HTMLResponse("No tienes permiso para ver esta etiqueta", status_code=403)

    cur.execute("SELECT COUNT(*) AS n FROM instrumentos WHERE envio_id=?", (envio_id,))
    n_inst = int(cur.fetchone()["n"] or 0)
    
    # -- Datos extra para Optica Rigida --
    fabricante, modelo, serie = "", "", ""
    tipo = str(e["tipo_trabajo"]).upper() if "tipo_trabajo" in e.keys() else "REPARACION"
    
    if tipo == "OPTICA_RIGIDA" and n_inst > 0:
        cur.execute("SELECT fabricante, codigo_producto, num_serie FROM instrumentos WHERE envio_id=? LIMIT 1", (envio_id,))
        inst_row = cur.fetchone()
        if inst_row:
            fabricante = inst_row["fabricante"] or ""
            modelo = inst_row["codigo_producto"] or ""
            serie = inst_row["num_serie"] or ""

    conn.close()

    # --- Fecha para la etiqueta (dd/mm/aaaa) ---
    fecha_raw = (e["fecha"] or "").strip()
    if fecha_raw:
        try:
            dt = datetime.fromisoformat(fecha_raw)
            fecha = dt.strftime("%d/%m/%Y")
        except ValueError:
            fecha = fecha_raw
    else:
        fecha = datetime.now().strftime("%d/%m/%Y")

    pdf = _build_etiqueta_pdf(
        str(e["ot_num"]),
        str(e["cliente"] or ""),
        fecha,
        n_inst,
        referencia=str(e["nombre_archivo"] or ""),
        fabricante=fabricante,
        modelo=modelo,
        serie=serie
    )
    filename = f"OT_{e['ot_num']}_etiqueta.pdf"
    return StreamingResponse(
        io.BytesIO(pdf),
        media_type="application/pdf",
        headers={"Content-Disposition": f"inline; filename={filename}"},
    )




# -----------------------------
# GRABACION (mosaico por parte)
# -----------------------------
@app.get("/envios/{envio_id}/grabacion", response_class=HTMLResponse)
def grabacion_envio(request: Request, envio_id: int, user=Depends(require_roles("admin", "grabado"))):
    conn = get_conn()
    cur = conn.cursor()

    cur.execute("SELECT * FROM envios WHERE id=?", (envio_id,))
    envio = cur.fetchone()
    if not envio:
        conn.close()
        return HTMLResponse("Envío no encontrado", status_code=404)

    cur.execute(
        """
        SELECT
          i.*,
          COALESCE(i.grabado,0) AS grabado
        FROM instrumentos i
        WHERE i.envio_id=?
        ORDER BY i.id DESC
        """,
        (envio_id,),
    )
    instrumentos = [dict(r) for r in cur.fetchall()]

    # Conteo
    total = len(instrumentos)
    grabados = sum(1 for r in instrumentos if int(r.get("grabado") or 0) == 1)

    conn.close()
    return templates.TemplateResponse(
        "envio_grabacion.html",
        {
            "request": request,
            "user": user,
            "envio": dict(envio),
            "instrumentos": instrumentos,
            "total": total,
            "grabados": grabados,
        },
    )


@app.post("/instrumentos/{instrumento_id}/grabar")
def grabar_instrumento(instrumento_id: int, user=Depends(require_roles("admin", "grabado"))):
    conn = get_conn()
    cur = conn.cursor()

    cur.execute("""
        SELECT i.envio_id, COALESCE(e.tipo_trabajo,'REPARACION') as tipo
        FROM instrumentos i
        JOIN envios e ON e.id = i.envio_id
        WHERE i.id=?
    """, (instrumento_id,))
    row = cur.fetchone()
    if not row:
        conn.close()
        return HTMLResponse("Instrumento no encontrado", status_code=404)

    envio_id = int(row["envio_id"])
    tipo_ot = (row["tipo"] or "").upper()

    if tipo_ot == "TRAZABILIDAD":
        cur.execute(
            """
            UPDATE instrumentos
            SET grabado=1,
                grabado_por=?,
                grabado_en=CURRENT_TIMESTAMP,
                estado='Reparado'
            WHERE id=?
            """,
            (int(user["id"]), instrumento_id),
        )
    else:
        cur.execute(
            """
            UPDATE instrumentos
            SET grabado=1,
                grabado_por=?,
                grabado_en=CURRENT_TIMESTAMP
            WHERE id=?
            """,
            (int(user["id"]), instrumento_id),
        )

    conn.commit()
    conn.close()

    return RedirectResponse(url=f"/envios/{envio_id}/grabacion", status_code=303)


@app.post("/instrumentos/{instrumento_id}/desgrabar")
def desgrabar_instrumento(instrumento_id: int, user=Depends(require_roles("admin"))):
    conn = get_conn()
    cur = conn.cursor()

    cur.execute("SELECT envio_id FROM instrumentos WHERE id=?", (instrumento_id,))
    row = cur.fetchone()
    if not row:
        conn.close()
        return HTMLResponse("Instrumento no encontrado", status_code=404)

    envio_id = int(row["envio_id"])

    cur.execute(
        """
        UPDATE instrumentos
        SET grabado=0,
            grabado_por=NULL,
            grabado_en=NULL
        WHERE id=?
        """,
        (instrumento_id,),
    )

    conn.commit()
    conn.close()

    return RedirectResponse(url=f"/envios/{envio_id}/grabacion", status_code=303)


# -----------------------------
# REVISIÓN FINAL (recepción)
# -----------------------------
@app.get("/envios/{envio_id}/revision", response_class=HTMLResponse)
def revision_envio(request: Request, envio_id: int, user=Depends(require_roles("admin", "recepcion"))):
    conn = get_conn()
    cur = conn.cursor()

    cur.execute("SELECT * FROM envios WHERE id=?", (envio_id,))
    envio = cur.fetchone()
    if not envio:
        conn.close()
        return HTMLResponse("Envío no encontrado", status_code=404)

    cur.execute(
        """
        SELECT i.*, 
               COALESCE(i.revisado,0) AS revisado,
               (SELECT 1 FROM instrumento_checklist ic WHERE ic.instrumento_id = i.id LIMIT 1) as has_checklist
        FROM instrumentos i
        WHERE i.envio_id=?
        ORDER BY i.id ASC
        """,
        (envio_id,),
    )
    instrumentos = []
    for r in cur.fetchall():
        d = dict(r)
        for k, v in d.items():
            if hasattr(v, "isoformat"): # Más genérico para datetime/date
                d[k] = v.isoformat()
        instrumentos.append(d)
    
    # Conteo
    n_revisados = sum(1 for i in instrumentos if i["revisado"])
    total = len(instrumentos)

    conn.close()
    return templates.TemplateResponse(
        "envio_revision.html",
        {
            "request": request,
            "user": user,
            "envio": dict(envio),
            "instrumentos": instrumentos,
            "total": total,
            "n_revisados": n_revisados,
        },
    )

@app.post("/instrumentos/{instrumento_id}/revisar")
def revisar_instrumento(instrumento_id: int, user=Depends(require_roles("admin", "recepcion"))):
    conn = get_conn()
    cur = conn.cursor()
    
    username = user.get("username") if isinstance(user, dict) else getattr(user, "username", "S/N")

    cur.execute("UPDATE instrumentos SET revisado=1, revisado_por=?, revisado_en=CURRENT_TIMESTAMP WHERE id=?", 
                (username, instrumento_id))
    conn.commit()
    conn.close()
    return {"ok": True}

@app.post("/instrumentos/{instrumento_id}/desrevisar")
def desrevisar_instrumento(instrumento_id: int, user=Depends(require_roles("admin", "recepcion"))):
    conn = get_conn()
    cur = conn.cursor()

    cur.execute("UPDATE instrumentos SET revisado=0, revisado_por=NULL, revisado_en=NULL WHERE id=?", 
                (instrumento_id,))
    conn.commit()
    conn.close()
    return {"ok": True}

# -----------------------------
# NUEVO INSTRUMENTO (admin/recepcion)
# -----------------------------
@app.get("/envios/{envio_id}/instrumentos/nuevo", response_class=HTMLResponse)
def instrumento_nuevo_form(request: Request, envio_id: int, user=Depends(require_roles("admin", "recepcion"))):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("SELECT * FROM envios WHERE id=?", (envio_id,))
    envio = cur.fetchone()
    conn.close()
    if not envio:
        return HTMLResponse("Envío no encontrado", status_code=404)

    return templates.TemplateResponse(
        "instrumento_nuevo.html",
        {
            "request": request,
            "user": user,
            "mode": "new",
            "envio": dict(envio),
            "inst": None,
        },
    )


@app.post("/envios/{envio_id}/instrumentos/nuevo")
def instrumento_nuevo_crear(
    envio_id: int,
    codigo_producto: str = Form(""),
    fabricante: str = Form(""),
    num_serie: str = Form(""),
    denominacion: str = Form(""),
    observaciones: str = Form(""),
    codigo_datamatrix: str = Form(""),
    codigo_cliente: str = Form(""),
    unidades: int = Form(1),
    user=Depends(require_roles("admin", "recepcion")),
):
    # Permiso granular (además del rol)
    conn_perm = get_conn()
    cur_perm = conn_perm.cursor()
    if not can_action(user, "instrumento_crear", cur_perm):
        conn_perm.close()
        return RedirectResponse(url="/?err=perm", status_code=303)
    conn_perm.close()

    conn = get_conn()
    cur = conn.cursor()

    cur.execute("SELECT id, cliente_id, tipo_trabajo FROM envios WHERE id=?", (envio_id,))
    e = cur.fetchone()
    if not e:
        conn.close()
        return HTMLResponse("Envío no encontrado", status_code=404)

    unidades = max(1, unidades)
    is_trazabilidad = (e["tipo_trabajo"] or "").upper() == "TRAZABILIDAD"
    
    prefijo_dm = ""
    prefijo_nombre = ""
    nums = []
    
    if is_trazabilidad:
        if not e["cliente_id"]:
            conn.close()
            return HTMLResponse("OT de trazabilidad sin cliente registrado", status_code=400)
        # Reservamos N números de golpe
        prefijo_dm, prefijo_nombre, nums = _reserve_numeros_cliente(cur, int(e["cliente_id"]), unidades)

    is_pg = bool(os.environ.get("DATABASE_URL"))
    inst_ids = []

    for i in range(unidades):
        dm_auto = ""
        nombre_trz_auto = ""
        if is_trazabilidad:
            dm_auto = f"{prefijo_dm}{str(nums[i]).zfill(5)}"
            nombre_trz_auto = _build_nombre_trazabilidad(prefijo_nombre, dm_auto)

        vals = [
            envio_id,
            (codigo_producto or "").strip(),
            (fabricante or "").strip(),
            (num_serie or "").strip(),
            (denominacion or "").strip(),
            (observaciones or "").strip(),
            (dm_auto or (codigo_datamatrix or "").strip()),
            (nombre_trz_auto or ""),
            (codigo_cliente or "").strip(), # NEW FIELD
        ]

        sql = """
            INSERT INTO instrumentos
            (envio_id, codigo_producto, fabricante, num_serie, denominacion, observaciones, codigo_datamatrix, nombre_trazabilidad, codigo_cliente, estado, creado_en)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, 'Pendiente', CURRENT_TIMESTAMP)
        """
        
        inst_id = None
        if is_pg:
            cur.execute(sql + " RETURNING id", tuple(vals))
            row = cur.fetchone()
            if row: inst_id = int(row["id"])
        else:
            cur.execute(sql, tuple(vals))
            inst_id = cur.lastrowid
            
        if inst_id is None:
            cur.execute("SELECT MAX(id) as mid FROM instrumentos WHERE envio_id=?", (envio_id,))
            row_id = cur.fetchone()
            if row_id and row_id["mid"]: inst_id = row_id["mid"]
        
        inst_ids.append(inst_id)

    conn.commit()
    conn.close()

    if unidades > 1:
        # Si son varias, volvemos a la lista del parte
        return RedirectResponse(url=f"/envios/{envio_id}", status_code=303)
    
    # Si es solo una, seguimos comportamiento original (ir a fotos/detalle)
    target_id = inst_ids[0]
    if is_trazabilidad:
        return RedirectResponse(url=f"/instrumentos/{target_id}", status_code=303)
    return RedirectResponse(url=f"/instrumentos/{target_id}/editar", status_code=303)


# -----------------------------
# EDITAR INSTRUMENTO (admin/recepcion)
# -----------------------------
@app.get("/instrumentos/{instrumento_id}/editar", response_class=HTMLResponse)
def instrumento_editar_form(request: Request, instrumento_id: int, user=Depends(require_roles("admin", "recepcion", "tecnico"))):
    conn = get_conn()
    cur = conn.cursor()

    cur.execute("SELECT * FROM instrumentos WHERE id=?", (instrumento_id,))
    inst = cur.fetchone()
    if not inst:
        conn.close()
        return HTMLResponse("Instrumento no encontrado", status_code=404)

    cur.execute("SELECT * FROM envios WHERE id=?", (inst["envio_id"],))
    envio = cur.fetchone()
    conn.close()

    return templates.TemplateResponse(
        "instrumento_nuevo.html",
        {
            "request": request,
            "user": user,
            "mode": "edit",
            "envio": dict(envio) if envio else None,
            "inst": dict(inst),
        },
    )


@app.post("/instrumentos/{instrumento_id}/editar")
def instrumento_editar_guardar(
    instrumento_id: int,
    codigo_producto: str = Form(""),
    fabricante: str = Form(""),
    num_serie: str = Form(""),
    denominacion: str = Form(""),
    observaciones: str = Form(""),
    codigo_datamatrix: str = Form(""),
    codigo_cliente: str = Form(""), # NEW FIELD
    user=Depends(require_roles("admin", "recepcion", "tecnico")),
):
    # Permiso granular
    conn_perm = get_conn()
    cur_perm = conn_perm.cursor()
    if not (can_action(user, "instrumento_editar", cur_perm) or can_action(user, "fotos_gestionar", cur_perm)):
        conn_perm.close()
        return RedirectResponse(url="/?err=perm", status_code=303)
    conn_perm.close()

    dm = (codigo_datamatrix or "").strip()

    conn = get_conn()
    cur = conn.cursor()

    # Obtener OT/cliente para saber si es trazabilidad y su prefijo_nombre
    cur.execute(
        """
        SELECT e.id AS envio_id, COALESCE(e.tipo_trabajo,'REPARACION') AS tipo_trabajo, e.cliente_id,
               c.prefijo_nombre
        FROM instrumentos i
        JOIN envios e ON e.id = i.envio_id
        LEFT JOIN clientes c ON c.id = e.cliente_id
        WHERE i.id=?
        """,
        (instrumento_id,),
    )
    meta = cur.fetchone()
    if not meta:
        conn.close()
        return HTMLResponse("Instrumento no encontrado", status_code=404)

    meta = dict(meta)

    nombre_trz = ""
    if (meta["tipo_trabajo"] or "").upper() == "TRAZABILIDAD":
        nombre_trz = _build_nombre_trazabilidad((meta.get("prefijo_nombre") or ""), dm)

    cur.execute(
        """
        UPDATE instrumentos
        SET codigo_producto=?,
            fabricante=?,
            num_serie=?,
            denominacion=?,
            observaciones=?,
            codigo_datamatrix=?,
            nombre_trazabilidad=?,
            codigo_cliente=?,
            actualizado_en=CURRENT_TIMESTAMP
        WHERE id=?
        """,
        (
            (codigo_producto or "").strip(),
            (fabricante or "").strip(),
            (num_serie or "").strip(),
            (denominacion or "").strip(),
            (observaciones or "").strip(),
            dm,
            nombre_trz,
            (codigo_cliente or "").strip(), # NEW FIELD
            instrumento_id,
        ),
    )

    conn.commit()
    conn.close()

    return RedirectResponse(url=f"/instrumentos/{instrumento_id}/editar", status_code=303)


@app.post("/instrumentos/{instrumento_id}/finalizar_fotos")
def instrumento_finalizar_fotos(instrumento_id: int, user=Depends(require_roles("admin", "recepcion", "tecnico"))):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("SELECT envio_id FROM instrumentos WHERE id=?", (instrumento_id,))
    row = cur.fetchone()
    conn.close()
    if not row:
        return HTMLResponse("Instrumento no encontrado", status_code=404)
    return RedirectResponse(url=f"/envios/{row['envio_id']}", status_code=303)


# -----------------------------
# ABRIR INSTRUMENTO (todos)
# -----------------------------
@app.get("/instrumentos/{instrumento_id}", response_class=HTMLResponse)
def instrumento_detalle(request: Request, instrumento_id: int, user=Depends(get_current_user)):
    conn = get_conn()
    cur = conn.cursor()

    cur.execute("""
        SELECT i.*, 
               (SELECT 1 FROM instrumento_checklist ic WHERE ic.instrumento_id = i.id LIMIT 1) as has_checklist
        FROM instrumentos i 
        WHERE i.id=?
    """, (instrumento_id,))
    inst = cur.fetchone()
    if not inst:
        conn.close()
        return HTMLResponse("Instrumento no encontrado", status_code=404)

    # Cargamos datos de la OT
    cur.execute("SELECT * FROM envios WHERE id=?", (inst["envio_id"],))
    envio = cur.fetchone()

    # SEGURIDAD: Si es rol 'cliente', verificamos pertenencia
    if _user_role(user) == "cliente":
        u_cli_id = (user.get("cliente_id") if isinstance(user, dict) else getattr(user, "cliente_id", None))
        if not envio or not u_cli_id or int(envio.get("cliente_id") or 0) != int(u_cli_id):
            conn.close()
            return HTMLResponse("Acceso denegado: este instrumento no pertenece a su centro", status_code=403)
    cliente = None
    if envio and envio["cliente_id"]:
        cur.execute("SELECT * FROM clientes WHERE id=?", (int(envio["cliente_id"]),))
        cliente = cur.fetchone()
    # Checklist configurable: por tipo de OT (REPARACION / OPTICA_RIGIDA / TRAZABILIDAD) y activo=1
    tipo_ot = "REPARACION"
    try:
        if envio is not None:
            # sqlite3.Row (o dict) -> tipo_trabajo
            if hasattr(envio, "keys") and ("tipo_trabajo" in envio.keys()):
                tipo_ot = envio["tipo_trabajo"] or "REPARACION"
            elif isinstance(envio, dict):
                tipo_ot = envio.get("tipo_trabajo") or "REPARACION"
    except Exception:
        tipo_ot = "REPARACION"
    tipo_ot = str(tipo_ot or "REPARACION").strip().upper()

    # Si es TRAZABILIDAD, usamos el checklist de REPARACION
    if tipo_ot == "TRAZABILIDAD":
        tipo_ot = "REPARACION"

    def _load_checklist(tipo: str):
        cur.execute("""
            SELECT ci.id AS item_id, ci.nombre, ci.orden,
                   COALESCE(ic.hecho,0) AS hecho,
                   ic.hecho_por, ic.hecho_en
            FROM checklist_items ci
            LEFT JOIN instrumento_checklist ic
              ON ic.item_id = ci.id AND ic.instrumento_id = ?
            WHERE COALESCE(ci.activo,1)=1
              AND COALESCE(ci.tipo_trabajo,'REPARACION') = ?
            ORDER BY LOWER(ci.nombre) ASC
        """, (instrumento_id, tipo))
        return [dict(r) for r in cur.fetchall()]

    checklist = _load_checklist(tipo_ot)

    # Fallback: si no hay checklist específico para ese tipo, usa el de REPARACION
    if (not checklist) and tipo_ot != "REPARACION":
        checklist = _load_checklist("REPARACION")

    # Informe PDF (óptica rígida): opcional
    cur.execute("SELECT * FROM instrumento_informes WHERE instrumento_id=? ORDER BY id DESC LIMIT 1", (instrumento_id,))
    informe = cur.fetchone()

    # --- HISTORIAL DE TRAZABILIDAD ---
    historial = []
    dm = (inst["codigo_datamatrix"] or "").strip()
    sn = (inst["num_serie"] or "").strip()
    if dm or sn:
        clauses = []
        p_hist = []
        if dm:
            clauses.append("i.codigo_datamatrix = ?")
            p_hist.append(dm)
        if sn:
            clauses.append("i.num_serie = ?")
            p_hist.append(sn)
        
        sql_hist = f"""
            SELECT i.id, i.envio_id, e.ot_num, e.fecha, i.estado, i.creado_en
            FROM instrumentos i
            JOIN envios e ON e.id = i.envio_id
            WHERE ({ " OR ".join(clauses) }) AND i.id != ?
        """
        p_hist.append(instrumento_id)

        if _user_role(user) == "cliente":
            u_cli_id = (user.get("cliente_id") if isinstance(user, dict) else getattr(user, "cliente_id", None))
            if u_cli_id:
                sql_hist += " AND e.cliente_id = ?"
                p_hist.append(int(u_cli_id))

        sql_hist += " ORDER BY e.id DESC LIMIT 10"
        cur.execute(sql_hist, tuple(p_hist))
        historial = [dict(r) for r in cur.fetchall()]

    # --- CHECKLIST DE REPUESTOS ---
    cur.execute("""
        SELECT rc.*, 
               COALESCE((SELECT ir.cantidad FROM instrumento_repuestos ir WHERE ir.instrumento_id=? AND ir.repuesto_id=rc.id), 0) as cantidad
        FROM repuestos_catalogo rc
        WHERE COALESCE(rc.activo,1)=1
        ORDER BY rc.nombre
    """, (instrumento_id,))
    repuestos_frecuentes = [dict(r) for r in cur.fetchall()]

    conn.close()

    return templates.TemplateResponse(
        "instrumento_detalle.html",
        {
            "request": request,
            "user": user,
            "inst": dict(inst),
            "checklist": checklist,
            "repuestos_frecuentes": repuestos_frecuentes,
            "envio": dict(envio) if envio else None,
            "cliente": dict(cliente) if cliente else None,
            "informe": dict(informe) if informe else None,
            "historial": historial
        },
    )
# -----------------------------
# CONTROL DE CALIDAD - OPTICAS RIGIDAS
# -----------------------------
@app.get("/instrumentos/{instrumento_id}/qc_optica", response_class=HTMLResponse)
async def qc_optica_view(request: Request, instrumento_id: int, user=Depends(require_roles("admin", "tecnico"))):
    conn = get_conn()
    cur = conn.cursor()
    
    cur.execute("SELECT * FROM instrumentos WHERE id=?", (instrumento_id,))
    inst = cur.fetchone()
    if not inst:
        conn.close()
        return HTMLResponse("Instrumento no encontrado", status_code=404)
        
    cur.execute("SELECT * FROM envios WHERE id=?", (inst["envio_id"],))
    envio = cur.fetchone()
    
    cur.execute("SELECT * FROM clientes WHERE id=?", (envio["cliente_id"],))
    cliente = cur.fetchone()
    
    cur.execute("SELECT * FROM instrumento_qc_optica WHERE instrumento_id=?", (instrumento_id,))
    qc = cur.fetchone()
    
    conn.close()
    
    return templates.TemplateResponse(
        "instrumento_qc_optica.html",
        {
            "request": request,
            "user": user,
            "inst": dict(inst),
            "envio": dict(envio),
            "cliente": dict(cliente) if cliente else None,
            "qc": dict(qc) if qc else None,
            "now_date": datetime.now().strftime("%Y-%m-%d")
        }
    )

@app.post("/instrumentos/{instrumento_id}/qc_optica")
async def qc_optica_save(
    instrumento_id: int,
    request: Request,
    user=Depends(require_roles("admin", "tecnico")),
    parte_trabajo_cliente: str = Form(""),
    observaciones_cliente: str = Form(""),
    observaciones_previas: str = Form(""),
    diag_ventana: str = Form("CORRECTO"),
    diag_fibra: str = Form("CORRECTO"),
    diag_objetivo: str = Form("CORRECTO"),
    diag_lentes: str = Form("CORRECTO"),
    diag_camisa: str = Form("CORRECTO"),
    diag_ocular: str = Form("CORRECTO"),
    diag_pieza_ojo: str = Form("CORRECTO"),
    diag_contaminacion: str = Form("CORRECTO"),
    reparable: int = Form(1),
    campo_vision_val: str = Form(""),
    campo_vision_ok: int = Form(1),
    direccion_vision_val: str = Form(""),
    direccion_vision_ok: int = Form(1),
    resolucion_val: str = Form(""),
    resolucion_ok: int = Form(1),
    desviacion_val: str = Form(""),
    desviacion_ok: int = Form(1),
    luz_val: str = Form(""),
    luz_ok: int = Form(1),
    observaciones_finales: str = Form(""),
    fecha_salida: str = Form(""),
    firma_tecnico: str = Form(""),
    firma_responsable: str = Form(""),
    qc_foto_entrada_1: UploadFile = File(None),
    qc_foto_entrada_2: UploadFile = File(None),
    qc_foto_salida_1: UploadFile = File(None),
    qc_foto_salida_2: UploadFile = File(None),
):
    conn = get_conn()
    cur = conn.cursor()
    
    # Manejo de fotos QC (entrada y salida)
    qc_fotos_vals = {}
    for key, f in [
        ("qc_foto_entrada_1", qc_foto_entrada_1), 
        ("qc_foto_entrada_2", qc_foto_entrada_2),
        ("qc_foto_salida_1", qc_foto_salida_1),
        ("qc_foto_salida_2", qc_foto_salida_2)
    ]:
        if f and f.filename:
            safe_name = re.sub(r"[^a-zA-Z0-9.-]", "_", f.filename)
            prefix = key.replace("qc_", "") # entrada_1, etc.
            fname = f"inst_{instrumento_id}_qc_{prefix}_{int(time.time())}_{safe_name}"
            path = os.path.join(FOTOS_DIR, fname)
            content = await f.read()
            
            # Intentar optimizar con Pillow, si falla guardar original
            try:
                from PIL import Image as PILImage
                import io
                
                img = PILImage.open(io.BytesIO(content))
                
                # Reducir si es muy grande (máx 1280px)
                max_size = 1290
                if img.width > max_size or img.height > max_size:
                    img.thumbnail((max_size, max_size), PILImage.LANCZOS)
                    
                # Convertir a RGB si es necesario (por si suben PNG con transp)
                if img.mode in ("RGBA", "P"):
                    img = img.convert("RGB")
                    
                # Guardar optimizada
                with open(path, "wb") as out:
                    img.save(out, format="JPEG", quality=85, optimize=True)
            except Exception as e:
                print(f"Error optimizando imagen {fname}: {e}")
                # Fallback: guardar tal cual
                with open(path, "wb") as out:
                    out.write(content)
            
            public_url = f"/static/fotos/{fname}"
            qc_fotos_vals[key] = public_url
            
            # Borrar anterior si existía en la tabla QC
            cur.execute(f"SELECT {key} FROM instrumento_qc_optica WHERE instrumento_id=?", (instrumento_id,))
            old = cur.fetchone()
            if old and old[key]:
                _try_delete_public_photo(old[key])

    # Guardar/Actualizar QC
    # Leemos todos los campos del form para manejar los dinámicos (split diagnosis)
    form_data = await request.form()
    
    cur.execute("SELECT 1 FROM instrumento_qc_optica WHERE instrumento_id=?", (instrumento_id,))
    exists = cur.fetchone()
    
    diag_items = ['ventana', 'fibra', 'objetivo', 'lentes', 'camisa', 'ocular', 'pieza_ojo', 'contaminacion']
    
    if exists:
        update_cols = [
            "parte_trabajo_cliente=?", "observaciones_cliente=?", "observaciones_previas=?",
            "reparable=?", "campo_vision_val=?", "campo_vision_ok=?", "direccion_vision_val=?", 
            "direccion_vision_ok=?", "resolucion_val=?", "resolucion_ok=?", "desviacion_val=?", 
            "desviacion_ok=?", "luz_val=?", "luz_ok=?", "observaciones_finales=?", "fecha_salida=?",
            "firma_tecnico=?", "firma_responsable=?"
        ]
        params = [
            parte_trabajo_cliente, observaciones_cliente, observaciones_previas,
            reparable, campo_vision_val, campo_vision_ok, direccion_vision_val, 
            direccion_vision_ok, resolucion_val, resolucion_ok, desviacion_val, 
            desviacion_ok, luz_val, luz_ok, observaciones_finales, fecha_salida,
            firma_tecnico, firma_responsable
        ]
        
        # Nuevas columnas de diagnóstico split
        for item in diag_items:
            update_cols.append(f"diag_{item}_estado=?")
            params.append(form_data.get(f"diag_{item}_estado", "CORRECTO"))
            update_cols.append(f"diag_{item}_accion=?")
            params.append(form_data.get(f"diag_{item}_accion", ""))
        
        for k, v in qc_fotos_vals.items():
            update_cols.append(f"{k}=?")
            params.append(v)
            
        params.append(instrumento_id)
        sql = f"UPDATE instrumento_qc_optica SET {', '.join(update_cols)} WHERE instrumento_id=?"
        cur.execute(sql, tuple(params))
    else:
        insert_cols = [
            "instrumento_id", "parte_trabajo_cliente", "observaciones_cliente", "observaciones_previas",
            "reparable", "campo_vision_val", "campo_vision_ok", "direccion_vision_val", "direccion_vision_ok",
            "resolucion_val", "resolucion_ok", "desviacion_val", "desviacion_ok",
            "luz_val", "luz_ok", "observaciones_finales", "fecha_salida",
            "firma_tecnico", "firma_responsable"
        ]
        params = [
            instrumento_id, parte_trabajo_cliente, observaciones_cliente, observaciones_previas,
            reparable, campo_vision_val, campo_vision_ok, direccion_vision_val, direccion_vision_ok,
            resolucion_val, resolucion_ok, desviacion_val, desviacion_ok,
            luz_val, luz_ok, observaciones_finales, fecha_salida,
            firma_tecnico, firma_responsable
        ]
        
        for item in diag_items:
            insert_cols.append(f"diag_{item}_estado")
            params.append(form_data.get(f"diag_{item}_estado", "CORRECTO"))
            insert_cols.append(f"diag_{item}_accion")
            params.append(form_data.get(f"diag_{item}_accion", ""))

        for k, v in qc_fotos_vals.items():
            insert_cols.append(k)
            params.append(v)
            
        placeholders = ", ".join(["?"] * len(insert_cols))
        sql = f"INSERT INTO instrumento_qc_optica ({', '.join(insert_cols)}) VALUES ({placeholders})"
        cur.execute(sql, tuple(params))

    conn.commit()
    conn.close()
    
    # --- NUEVOS PASOS PARA ARCHIVAR EL PDF AUTOMÁTICAMENTE ---
    conn = get_conn()
    
    try:
        cur = conn.cursor()
        # Recargar datos frescos para el generador
        pdf_bytes, filename = _generate_qc_optica_pdf_bytes(instrumento_id)
        if pdf_bytes:
            informes_dir = os.path.join(UPLOAD_DIR, "informes")
            os.makedirs(informes_dir, exist_ok=True)
            # Nombre único con timestamp para no machacar registros históricos si se quiere
            stored_name = f"auto_qc_{instrumento_id}_{int(time.time())}.pdf"
            full_path = os.path.join(informes_dir, stored_name)
            with open(full_path, "wb") as f_pdf:
                f_pdf.write(pdf_bytes)
                
            # Almacenar la ruta relativa para que sea accesible vía web
            db_path = f"uploads/informes/{stored_name}"
            
            # Insertar en instrumento_informes para que aparezca en el listado
            from db import get_table_columns
            cols = get_table_columns(cur, "instrumento_informes")
            now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            
            username = (user.get("username") if isinstance(user, dict) else getattr(user, "username", "SISTEMA"))
            
            insert_cols = ["instrumento_id", "filename"]
            insert_vals = [instrumento_id, filename]
            if "path" in cols:
                insert_cols.append("path")
                insert_vals.append(db_path)
            if "filepath" in cols:
                insert_cols.append("filepath")
                insert_vals.append(db_path)
            if "uploaded_at" in cols:
                insert_cols.append("uploaded_at")
                insert_vals.append(now_str)
            if "uploaded_by" in cols:
                insert_cols.append("uploaded_by")
                insert_vals.append(username)
                
            placeholders = ", ".join(["?"] * len(insert_cols))
            cur.execute(f"INSERT INTO instrumento_informes ({', '.join(insert_cols)}) VALUES ({placeholders})", tuple(insert_vals))
            conn.commit()
    except Exception as e:
        print(f"Error generando PDF automatico: {e}")
        # Continuamos con el resto (marcar reparado etc)
        
    # Marcar instrumento como REPARADO al guardar el QC
    try:
        cur = conn.cursor()
        cur.execute("UPDATE instrumentos SET estado='Reparado' WHERE id=?", (instrumento_id,))
        conn.commit()

        # Redirigir de vuelta al listado del parte (envio_id)
        cur.execute("SELECT envio_id FROM instrumentos WHERE id=?", (instrumento_id,))
        row_e = cur.fetchone()
    except Exception as e:
        print(f"Error actualizando estado o redireccionando: {e}")
        row_e = None
    finally:
        conn.close()
    
    dest_url = f"/envios/{row_e['envio_id']}" if row_e else "/"
    return RedirectResponse(url=dest_url, status_code=303)


def _generate_qc_optica_pdf_bytes(instrumento_id: int):
    conn = get_conn()
    cur = conn.cursor()
    
    cur.execute("SELECT * FROM instrumentos WHERE id=?", (instrumento_id,))
    inst = cur.fetchone()
    if not inst:
        conn.close()
        return None, None
        
    cur.execute("SELECT * FROM envios WHERE id=?", (inst["envio_id"],))
    envio = cur.fetchone()
    
    cur.execute("SELECT * FROM clientes WHERE id=?", (envio["cliente_id"],))
    cliente = cur.fetchone()
    
    cur.execute("SELECT * FROM instrumento_qc_optica WHERE instrumento_id=?", (instrumento_id,))
    qc_row = cur.fetchone()
    conn.close()
    
    if not qc_row:
        return None, None

    # Convertir a dict para poder usar .get() y evitar errores de sqlite3.Row
    inst = dict(inst)
    envio = dict(envio)
    cliente = dict(cliente) if cliente else None
    qc = dict(qc_row)

    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image, HRFlowable, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, rightMargin=1.5*cm, leftMargin=1.5*cm, topMargin=1.5*cm, bottomMargin=1.5*cm)
    styles = getSampleStyleSheet()
    
    # Colores Corporativos
    c_navy = colors.HexColor('#002D62')
    c_light_bg = colors.HexColor('#F4F7F9')
    c_border = colors.HexColor('#D1D5DB')
    
    # Estilos Personalizados
    style_h1 = ParagraphStyle('H1', parent=styles['Heading1'], fontSize=16, textColor=c_navy, alignment=TA_CENTER, spaceAfter=20, fontName='Helvetica-Bold')
    style_sec = ParagraphStyle('Section', parent=styles['Heading2'], fontSize=11, textColor=colors.black, fontName='Helvetica-Bold', spaceBefore=10, spaceAfter=8, borderPadding=2)
    style_normal = ParagraphStyle('NormalSmall', parent=styles['Normal'], fontSize=9, leading=11)
    style_cell_label = ParagraphStyle('CellLabel', parent=styles['Normal'], fontSize=8, fontName='Helvetica-Bold', textColor=c_navy)

    elements = []

    # --- CABECERA ---
    logo_path = os.path.join("static", "logo-xurgical.png")
    logo = Image(logo_path, width=4.5*cm, height=1.35*cm) if os.path.exists(logo_path) else Paragraph("XURGICAL", styles["Normal"])
    
    header_right = [
        [Paragraph("<b>CERTIFICADO DE CONTROL DE CALIDAD</b>", ParagraphStyle('cc', fontSize=12, alignment=TA_RIGHT, textColor=c_navy))],
        [Paragraph(f"Informe Nº: CC-{24000 + instrumento_id}", ParagraphStyle('id', fontSize=10, alignment=TA_RIGHT))],
        [Paragraph(f"Fecha: {qc['fecha_salida'] or datetime.now().strftime('%d/%m/%Y')}", ParagraphStyle('dt', fontSize=10, alignment=TA_RIGHT))]
    ]
    header_tab = Table([[logo, Table(header_right, colWidths=[8*cm])]], colWidths=[8*cm, 10*cm])
    header_tab.setStyle(TableStyle([('VALIGN', (0,0), (-1,-1), 'MIDDLE'), ('ALIGN', (1,0), (1,0), 'RIGHT')]))
    elements.append(header_tab)
    elements.append(HRFlowable(width="100%", thickness=1, color=c_navy, spaceBefore=4, spaceAfter=15))

    # --- DATOS GENERALES ---
    elements.append(Paragraph("DATOS DEL EQUIPO Y CLIENTE", style_sec))
    
    dg_data = [
        [Paragraph("CLIENTE", style_cell_label), str(cliente["nombre"] if cliente else envio["cliente"])[:50]],
        [Paragraph("DENOMINACIÓN", style_cell_label), str(inst["denominacion"])[:50]],
        [Paragraph("MODELO / REF", style_cell_label), str(inst["codigo_producto"])[:30]],
        [Paragraph("Nº DE SERIE", style_cell_label), str(inst["num_serie"])[:30]],
        [Paragraph("ORDEN TRABAJO", style_cell_label), str(envio["ot_num"])[:20]],
        ["", ""]
    ]
    # Reorganizar en 2 columnas
    dg_tab_inner = [
        [dg_data[0][0], dg_data[0][1], dg_data[1][0], dg_data[1][1]],
        [dg_data[2][0], dg_data[2][1], dg_data[3][0], dg_data[3][1]],
        [dg_data[4][0], dg_data[4][1], "", ""]
    ]
    dg_tab = Table(dg_tab_inner, colWidths=[3.5*cm, 5.5*cm, 3.5*cm, 5.5*cm])
    dg_tab.setStyle(TableStyle([
        ('GRID', (0,0), (-1,-1), 0.5, c_border),
        ('BACKGROUND', (0,0), (0,-1), c_light_bg),
        ('BACKGROUND', (2,0), (2,-1), c_light_bg),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('FONTSIZE', (0,0), (-1,-1), 9),
        ('LEFTPADDING', (0,0), (-1,-1), 6),
    ]))
    elements.append(dg_tab)
    
    # Observaciones adicionales
    obs_extra_data = [
        [Paragraph("OBS. CLIENTE", style_cell_label), Paragraph(qc["observaciones_cliente"] or "-", style_normal)],
        [Paragraph("OBS. PREVIAS", style_cell_label), Paragraph(qc["observaciones_previas"] or "-", style_normal)]
    ]
    obs_extra_tab = Table(obs_extra_data, colWidths=[3.5*cm, 14.5*cm])
    obs_extra_tab.setStyle(TableStyle([
        ('GRID', (0,0), (-1,-1), 0.5, c_border),
        ('BACKGROUND', (0,0), (-1,-1), c_light_bg),
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('TOPPADDING', (0,0), (-1,-1), 4),
        ('BOTTOMPADDING', (0,0), (-1,-1), 4),
    ]))
    elements.append(obs_extra_tab)
    elements.append(Spacer(1, 15))

    # --- DIAGNÓSTICO TÉCNICO ---
    elements.append(Paragraph("INSPECCIÓN Y DIAGNÓSTICO DE COMPONENTES", style_sec))
    diag_rows = [
        [
            Paragraph("ELEMENTO", style_cell_label),
            Paragraph("EVALUACIÓN", style_cell_label),
            "",
            Paragraph("ACCIÓN (SI INCORRECTO)", style_cell_label),
            ""
        ],
        [
            "",
            Paragraph("CORR.", style_cell_label),
            Paragraph("INCOR.", style_cell_label),
            Paragraph("SUST.", style_cell_label),
            Paragraph("REPAR.", style_cell_label)
        ]
    ]
    
    elementos_qc = [
        ('ventana', 'VENTANA'), ('fibra', 'FIBRA ILUMINACIÓN'), ('objetivo', 'OBJETIVO'),
        ('lentes', 'LENTES'), ('camisa', 'CAMISA EXTERIOR'), ('ocular', 'OCULAR'),
        ('pieza_ojo', 'PIEZA DE OJO'), ('contaminacion', 'CONTAMINACIÓN')
    ]
    
    for item, label in elementos_qc:
        est = qc.get(f"diag_{item}_estado", "CORRECTO")
        acc = qc.get(f"diag_{item}_accion", "")
        diag_rows.append([
            label, 
            "X" if est=="CORRECTO" else "", 
            "X" if est=="INCORRECTO" else "", 
            "X" if acc=="SUSTITUCION" else "", 
            "X" if acc=="REPARACION" else ""
        ])
        
    diag_tab = Table(diag_rows, colWidths=[6*cm, 3*cm, 3*cm, 3*cm, 3*cm])
    diag_tab.setStyle(TableStyle([
        ('GRID', (0,0), (-1,-1), 0.5, c_border),
        ('SPAN', (1,0), (2,0)), # Span Evaluación
        ('SPAN', (3,0), (4,0)), # Span Acción
        ('SPAN', (0,0), (0,1)), # Span Elemento header
        ('BACKGROUND', (0,0), (-1,1), c_light_bg),
        ('BACKGROUND', (3,0), (4,-1), colors.HexColor('#F9FAFB')), # Fondo sutil para acción
        ('FONTSIZE', (0,0), (-1,-1), 8),
        ('ALIGN', (1,1), (-1,-1), 'CENTER'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('LEFTPADDING', (0,0), (-1,-1), 8),
    ]))
    elements.append(diag_tab)
    
    # Salto de página para que las fotos y lo demás vayan a la pág 2
    elements.append(PageBreak())

    # --- FOTOS ---
    elements.append(Paragraph("REGISTRO FOTOGRÁFICO DE DIAGNÓSTICO Y REPARACIÓN", style_sec))
    
    def _get_pdf_img(p):
        if not p: return Paragraph("<br/><br/><i>Sin imagen</i>", ParagraphStyle('si', alignment=TA_CENTER, fontSize=8, textColor=colors.gray))
        
        # Intentar extraer el nombre del archivo si es una URL/ruta
        fname = p.split("/")[-1]
        
        # 1. Probar en FOTOS_DIR (disco persistente o carpeta configurada)
        loc = os.path.join(FOTOS_DIR, fname)
        
        if not os.path.exists(loc):
            # 2. Probar en BASE_DIR/static/fotos (ruta por defecto de Reportlab en algunos entornos)
            loc = os.path.join(str(BASE_DIR), "static", "fotos", fname)
            
            if not os.path.exists(loc):
                # 3. Fallback: probar quitando el '/' de la ruta almacenada
                loc = p.lstrip("/")
                if not os.path.exists(loc):
                    # No se encuentra el archivo físico
                    return Paragraph("<br/><br/><i>Imagen no encontrada</i>", ParagraphStyle('err', alignment=TA_CENTER, fontSize=8, textColor=colors.gray))
        
        try:
            return Image(loc, width=7.5*cm, height=5.5*cm, kind='proportional')
        except Exception as e:
            print(f"Error cargando imagen PDF: {e}")
            return Paragraph("<br/><br/><i>Error de carga</i>", ParagraphStyle('err', alignment=TA_CENTER, fontSize=8, textColor=colors.gray))

    foto_data = [
        [Paragraph("<b>ESTADO INICIAL / ENTRADA</b>", style_normal), Paragraph("<b>ESTADO FINAL / SALIDA</b>", style_normal)],
        [_get_pdf_img(qc["qc_foto_entrada_1"] or inst["foto_entrada_1"]), _get_pdf_img(qc["qc_foto_salida_1"] or inst["foto_salida_1"])],
        [_get_pdf_img(qc["qc_foto_entrada_2"] or inst["foto_entrada_2"]), _get_pdf_img(qc["qc_foto_salida_2"] or inst["foto_salida_2"])]
    ]
    foto_tab = Table(foto_data, colWidths=[8.5*cm, 8.5*cm])
    foto_tab.setStyle(TableStyle([
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('GRID', (0,1), (-1,-1), 0.5, c_border),
        ('TOPPADDING', (0,0), (-1,-1), 5),
        ('BOTTOMPADDING', (0,0), (-1,-1), 5),
    ]))
    elements.append(foto_tab)

    # --- VERIFICACIÓN DE PARÁMETROS ---
    elements.append(Paragraph("VERIFICACIÓN DE ESPECIFICACIONES TÉCNICAS", style_sec))
    tec_rows = [[
        Paragraph("PARÁMETRO", style_cell_label),
        Paragraph("VALOR", style_cell_label),
        Paragraph("VÁLIDO (SÍ)", style_cell_label),
        Paragraph("VÁLIDO (NO)", style_cell_label)
    ]]
    for key, label in [('campo_vision', 'CAMPO DE VISIÓN'), ('direccion_vision', 'DIRECCIÓN DE VISIÓN'), 
                       ('resolucion', 'RESOLUCIÓN'), ('desviacion', 'DESVIACIÓN'), ('luz', 'LUZ')]:
        ok = qc[key+"_ok"]
        tec_rows.append([label, qc[key+"_val"] or "-", "X" if ok==1 else "", "X" if ok==0 else ""])
        
    tec_tab = Table(tec_rows, colWidths=[6.5*cm, 4.5*cm, 3.5*cm, 3.5*cm])
    tec_tab.setStyle(TableStyle([
        ('GRID', (0,0), (-1,-1), 0.5, c_border),
        ('BACKGROUND', (0,0), (-1,0), c_light_bg),
        ('FONTSIZE', (0,0), (-1,-1), 8),
        ('ALIGN', (1,1), (-1,-1), 'CENTER'),
        ('LEFTPADDING', (0,0), (-1,-1), 8),
    ]))
    elements.append(tec_tab)
    elements.append(Spacer(1, 15))

    # --- CIERRE ---
    elements.append(Paragraph("CONCLUSIONES Y OBSERVACIONES FINALES", style_sec))
    obs = qc["observaciones_finales"] or "El equipo ha sido sometido a las pruebas de control de calidad indicadas, resultando apto para su uso clínico tras la intervención realizada."
    elements.append(Paragraph(obs, style_normal))
    
    elements.append(Spacer(1, 20))
    
    # Firmas
    firma_data = [
        [Paragraph(f"<b>CONTROL DE CALIDAD</b><br/><br/><br/>{qc['firma_tecnico'] or 'Técnico Especialista'}", style_normal), 
         Paragraph(f"<b>RESPONSABLE TÉCNICO</b><br/><br/><br/>{qc['firma_responsable'] or 'Director Técnico'}", style_normal)]
    ]
    firma_tab = Table(firma_data, colWidths=[8.5*cm, 8.5*cm])
    firma_tab.setStyle(TableStyle([('ALIGN', (0,0), (-1,-1), 'CENTER')]))
    elements.append(firma_tab)

    # --- FOOTER (Se genera en cada página) ---
    def static_footer(canvas, doc):
        canvas.saveState()
        canvas.setFont('Helvetica', 8)
        canvas.setStrokeColor(c_navy)
        canvas.line(1.5*cm, 1.2*cm, 19.5*cm, 1.2*cm)
        footer_text = "XURGICAL - Soluciones en Instrumentación Quirúrgica | www.xurgical.com"
        canvas.drawCentredString(10.5*cm, 0.8*cm, footer_text)
        canvas.restoreState()

    doc.build(elements, onFirstPage=static_footer, onLaterPages=static_footer)
    pdf_bytes = buf.getvalue()
    clean_code = str(inst['codigo_producto'] or 'SREF').replace("/", "_").replace("\\", "_")
    filename = f"QC_OPTICA_{instrumento_id}_{clean_code}.pdf"
    return pdf_bytes, filename

@app.get("/instrumentos/{instrumento_id}/qc_optica/pdf")
async def qc_optica_pdf_gen(instrumento_id: int, user=Depends(get_current_user)):
    conn = get_conn()
    cur = conn.cursor()
    
    # Seguridad para clientes
    if _user_role(user) == "cliente":
        cur.execute("""
            SELECT e.cliente_id 
            FROM instrumentos i 
            JOIN envios e ON e.id = i.envio_id 
            WHERE i.id=?
        """, (instrumento_id,))
        row = cur.fetchone()
        u_cli_id = (user.get("cliente_id") if isinstance(user, dict) else getattr(user, "cliente_id", None))
        if not row or not u_cli_id or int(row["cliente_id"] or 0) != int(u_cli_id):
            conn.close()
            raise HTTPException(status_code=403, detail="Acceso denegado a este informe")
    
    conn.close()
    pdf_bytes, filename = _generate_qc_optica_pdf_bytes(instrumento_id)
    if not pdf_bytes:
        raise HTTPException(status_code=404, detail="No hay datos de QC para este instrumento")
    
    return StreamingResponse(
        io.BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f"inline; filename={filename}"}
    )


# -----------------------------
# CHECKLIST (admin/tecnico)
# Endpoint que usa el botón OK del checklist.
# Si no existe, el navegador devuelve {"detail":"No encontrado"}.
# -----------------------------
# -----------------------------
# INFORME PDF (óptica rígida)
# -----------------------------
@app.get("/instrumentos/{instrumento_id}/informe.pdf", response_class=FileResponse)
@app.get("/instrumentos/{instrumento_id}/informe", response_class=FileResponse)
def instrumento_informe_download(instrumento_id: int, user=Depends(get_current_user)):
    conn = get_conn()
    cur = conn.cursor()

    # Seguridad para clientes
    if _user_role(user) == "cliente":
        cur.execute("""
            SELECT e.cliente_id 
            FROM instrumentos i 
            JOIN envios e ON e.id = i.envio_id 
            WHERE i.id=?
        """, (instrumento_id,))
        row_sec = cur.fetchone()
        u_cli_id = (user.get("cliente_id") if isinstance(user, dict) else getattr(user, "cliente_id", None))
        if not row_sec or not u_cli_id or int(row_sec["cliente_id"] or 0) != int(u_cli_id):
            conn.close()
            raise HTTPException(status_code=403, detail="Acceso denegado a este informe")

    cur.execute("SELECT path, filename FROM instrumento_informes WHERE instrumento_id=? ORDER BY id DESC LIMIT 1", (instrumento_id,))
    row = cur.fetchone()
    conn.close()
    
    # Si existe el registro y el archivo en disco, lo servimos directamente
    if row and row["path"] and os.path.exists(row["path"]):
        return FileResponse(row["path"], filename=row["filename"], media_type="application/pdf")
    
    # FALLBACK INMINENTE: Si no hay archivo guardado, intentamos generarlo al vuelo
    pdf_bytes, filename = _generate_qc_optica_pdf_bytes(instrumento_id)
    if pdf_bytes:
        return StreamingResponse(
            io.BytesIO(pdf_bytes),
            media_type="application/pdf",
            headers={"Content-Disposition": f"inline; filename={filename}"}
        )
        
    raise HTTPException(status_code=404, detail="Informe no encontrado y no se puede generar (faltan datos QC)")


@app.post("/instrumentos/{instrumento_id}/informe/upload")
async def instrumento_informe_upload(
    instrumento_id: int,
    file: UploadFile = File(...),
    user=Depends(require_roles("admin", "recepcion", "tecnico")),
):
    """Sube un informe PDF para un instrumento.

    Esta versión es compatible con BDs antiguas que usaban la columna `filepath`
    (incluso si era NOT NULL). Inserta dinámicamente en `path` y/o `filepath`
    según existan en la tabla real.
    """
    if not (file.filename or "").lower().endswith(".pdf"):
        raise HTTPException(status_code=400, detail="Solo se permite PDF")

    informes_dir = os.path.join(UPLOAD_DIR, "informes")
    os.makedirs(informes_dir, exist_ok=True)

    safe_name = re.sub(r"[^a-zA-Z0-9._-]", "_", file.filename)
    stored_name = f"inst_{instrumento_id}_{int(time.time())}_{safe_name}"
    path = os.path.join(informes_dir, stored_name)

    with open(path, "wb") as f:
        f.write(await file.read())

    conn = get_conn()
    cur = conn.cursor()

    # Descubre columnas reales
    from db import get_table_columns
    cols = get_table_columns(cur, "instrumento_informes")

    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # user puede ser dict (tu auth) o un objeto
    username = None
    try:
        if isinstance(user, dict):
            username = user.get("username")
        else:
            username = getattr(user, "username", None)
    except Exception:
        username = None

    insert_cols = ["instrumento_id", "filename"]
    insert_vals = [instrumento_id, file.filename]

    if "path" in cols:
        insert_cols.append("path")
        insert_vals.append(path)

    if "filepath" in cols:
        insert_cols.append("filepath")
        insert_vals.append(path)

    if "uploaded_at" in cols:
        insert_cols.append("uploaded_at")
        insert_vals.append(now)

    if "uploaded_by" in cols:
        insert_cols.append("uploaded_by")
        insert_vals.append(username)

    if ("path" not in cols) and ("filepath" not in cols):
        conn.close()
        raise HTTPException(status_code=500, detail="La tabla instrumento_informes no tiene columnas path/filepath")

    placeholders = ", ".join(["?"] * len(insert_cols))
    sql = f"INSERT INTO instrumento_informes ({', '.join(insert_cols)}) VALUES ({placeholders})"
    cur.execute(sql, tuple(insert_vals))

    conn.commit()
    conn.close()

    return RedirectResponse(url=f"/instrumentos/{instrumento_id}", status_code=303)


@app.post("/instrumentos/{instrumento_id}/check/{item_id}")
async def checklist_toggle(
    instrumento_id: int,
    item_id: int,
    request: Request,
    user=Depends(require_roles("admin", "tecnico")),
):
    """Marca/desmarca un item del checklist para un instrumento.

    Admite POST desde formulario (por ejemplo, botón OK) y devuelve redirect al
    detalle del instrumento.

    - Si el form incluye `hecho` ("1"/"0"/"on"), se usa ese valor.
    - Si no incluye `hecho`, se hace toggle del estado actual.
    """

    conn = get_conn()
    cur = conn.cursor()

    # Verifica instrumento
    cur.execute("SELECT id FROM instrumentos WHERE id=?", (instrumento_id,))
    if not cur.fetchone():
        conn.close()
        raise HTTPException(status_code=404, detail="No encontrado")

    # Verifica item
    cur.execute("SELECT id FROM checklist_items WHERE id=?", (item_id,))
    if not cur.fetchone():
        conn.close()
        raise HTTPException(status_code=404, detail="No encontrado")

    form = await request.form()
    raw = form.get("hecho")

    # Estado actual
    cur.execute(
        "SELECT hecho FROM instrumento_checklist WHERE instrumento_id=? AND item_id=?",
        (instrumento_id, item_id),
    )
    row = cur.fetchone()
    actual = int(row["hecho"]) if row and isinstance(row, dict) else (int(row[0]) if row else 0)

    if raw is None:
        nuevo = 0 if actual == 1 else 1
    else:
        s = str(raw).strip().lower()
        nuevo = 1 if s in ("1", "true", "on", "yes", "si") else 0

    hecho_por = user.get("username") if isinstance(user, dict) else None

    if row:
        cur.execute(
            """
            UPDATE instrumento_checklist
               SET hecho=?, hecho_por=?, hecho_en=CURRENT_TIMESTAMP
             WHERE instrumento_id=? AND item_id=?
            """,
            (nuevo, hecho_por, instrumento_id, item_id),
        )
    else:
        cur.execute(
            """
            INSERT INTO instrumento_checklist (instrumento_id, item_id, hecho, hecho_por, hecho_en)
            VALUES (?, ?, ?, ?, CURRENT_TIMESTAMP)
            """,
            (instrumento_id, item_id, nuevo, hecho_por),
        )

    conn.commit()
    conn.close()

    # Si viene por AJAX, devolvemos JSON
    accept = (request.headers.get("accept") or "").lower()
    xreq = (request.headers.get("x-requested-with") or "").lower()
    if "application/json" in accept or xreq == "xmlhttprequest":
        return {"ok": True, "instrumento_id": instrumento_id, "item_id": item_id, "hecho": nuevo}

    return RedirectResponse(url=f"/instrumentos/{instrumento_id}", status_code=303)


# -----------------------------
# CAMBIAR ESTADO (admin/tecnico)
# -----------------------------
@app.post("/instrumentos/{instrumento_id}/estado")
async def cambiar_estado(
    instrumento_id: int,
    request: Request,
    user=Depends(require_roles("admin", "tecnico")),
):
    """Cambia el estado del instrumento (Pendiente/En proceso/Reparado/Baja) y vuelve a la página anterior."""
    form = await request.form()
    estado = (form.get("estado") or "").strip()

    if estado not in ("Pendiente", "En proceso", "Reparado", "Baja"):
        return HTMLResponse("Estado inválido", status_code=400)

    conn = get_conn()
    cur = conn.cursor()

    # Si la OT es de TRAZABILIDAD, no se permite cambiar estado de reparación.
    cur.execute("SELECT envio_id FROM instrumentos WHERE id=?", (instrumento_id,))
    row = cur.fetchone()
    envio_id = int(row["envio_id"]) if row and row["envio_id"] is not None else None

    if envio_id is not None and _envios_has_column(cur, "tipo_trabajo"):
        cur.execute("SELECT tipo_trabajo FROM envios WHERE id=?", (envio_id,))
        e = cur.fetchone()
        tipo = (e["tipo_trabajo"] if e and e["tipo_trabajo"] else "REPARACION")
        # if tipo == "TRAZABILIDAD":
        #     conn.close()
        #     return HTMLResponse("En OTs de trazabilidad no se cambia el estado de reparación.", status_code=400)

    # Actualiza estado y repuestos
    repuesto_info = form.get("repuesto_info")
    repuesto_precio_raw = form.get("repuesto_precio")
    repuesto_precio = None
    if repuesto_precio_raw:
        try:
            repuesto_precio = float(repuesto_precio_raw)
        except:
            repuesto_precio = None

    # Recomendación de sustitución
    sustitucion = 1 if form.get("recomendada_sustitucion") else 0

    ahora = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    username = user.get("username")

    cur.execute("""
        UPDATE instrumentos 
        SET estado=?, 
            tecnico_reparacion=?, 
            tecnico_reparacion_en=?,
            repuesto_info=?,
            repuesto_precio=?,
            recomendada_sustitucion=?
        WHERE id=?
    """, (estado, username, ahora, repuesto_info, repuesto_precio, sustitucion, instrumento_id))
    
    conn.commit()
    conn.close()

    if envio_id:
        return RedirectResponse(url=f"/envios/{envio_id}", status_code=303)
    return RedirectResponse(url="/", status_code=303)



# -----------------------------
# SUBIR FOTO DESDE WEBCAM (admin/recepcion)
# Body JSON: { image: "data:image/jpeg;base64,...." }
# -----------------------------
@app.post("/instrumentos/{instrumento_id}/foto_webcam/{slot}")
async def foto_webcam(instrumento_id: int, slot: int, request: Request, user=Depends(require_roles("admin", "recepcion", "tecnico"))):
    if slot not in range(1, 7):
        return JSONResponse({"ok": False, "error": "slot inválido"}, status_code=400)

    data = await request.json()
    image = (data.get("image") or "")
    if "," in image:
        image = image.split(",", 1)[1]

    try:
        raw = base64.b64decode(image)
    except Exception:
        return JSONResponse({"ok": False, "error": "imagen inválida"}, status_code=400)

    col = f"foto_entrada_{slot}"

    conn = get_conn()
    cur = conn.cursor()
    cur.execute(f"SELECT codigo_datamatrix, nombre_trazabilidad, {col} AS old FROM instrumentos WHERE id=?", (instrumento_id,))
    row = cur.fetchone()
    if not row:
        conn.close()
        return JSONResponse({"ok": False, "error": "instrumento no encontrado"}, status_code=404)

    row = dict(row)

    # nombre para la foto: prioriza nombre_trazabilidad (TRZ), si no usa datamatrix
    tag = (row.get("nombre_trazabilidad") or row.get("codigo_datamatrix") or "").strip()
    tag = re.sub(r"[^A-Za-z0-9_-]+", "_", tag)[:40] if tag else "SIN_CODIGO"

    filename = f"inst_{instrumento_id}_{tag}_f{slot}_{uuid.uuid4().hex[:8]}.jpg"
    path_fs = os.path.join(FOTOS_DIR, filename)
    print(f"DEBUG PHOTO: Guardando foto en {path_fs}")
    with open(path_fs, "wb") as f:
        f.write(raw)

    public_path = f"/static/fotos/{filename}"

    # borrar anterior si existe
    old_path = row.get("old")
    if old_path:
        _try_delete_public_photo(old_path)

    cur.execute(f"UPDATE instrumentos SET {col}=? WHERE id=?", (public_path, instrumento_id))
    conn.commit()
    conn.close()

    return {"ok": True, "path": public_path}



# -----------------------------
# CHECKLIST ADMIN (configurable)
# -----------------------------
@app.get("/checklist_admin", response_class=HTMLResponse)
def checklist_admin(request: Request, tipo: str = "REPARACION", user=Depends(require_roles("admin"))):
    tipo = (tipo or "REPARACION").strip().upper()
    if tipo not in ("REPARACION", "TRAZABILIDAD", "OPTICA_RIGIDA"):
        tipo = "REPARACION"

    conn = get_conn()
    cur = conn.cursor()
    cur.execute("""
        SELECT id, nombre, orden,
               COALESCE(activo,1) AS activo,
               COALESCE(tipo_trabajo,'REPARACION') AS tipo_trabajo
        FROM checklist_items
        WHERE COALESCE(tipo_trabajo,'REPARACION') = ?
        ORDER BY LOWER(nombre) ASC
    """, (tipo,))
    items = [dict(r) for r in cur.fetchall()]
    conn.close()

    return templates.TemplateResponse(
        "checklist_admin.html",
        {
            "request": request,
            "user": user,
            "tipo": tipo,
            "items": items,
            "tipos": ["REPARACION", "OPTICA_RIGIDA", "TRAZABILIDAD"],
        },
    )


@app.post("/checklist_admin/add")
async def checklist_admin_add(
    request: Request,
    user=Depends(require_roles("admin")),
):
    form = await request.form()
    nombre = (form.get("nombre") or "").strip()
    tipo = (form.get("tipo_trabajo") or "REPARACION").strip().upper()
    orden = form.get("orden") or "0"

    try:
        orden_i = int(str(orden).strip() or "0")
    except Exception:
        orden_i = 0

    if not nombre:
        return RedirectResponse(url=f"/checklist_admin?tipo={tipo}", status_code=303)

    if tipo not in ("REPARACION", "TRAZABILIDAD", "OPTICA_RIGIDA"):
        tipo = "REPARACION"

    conn = get_conn()
    cur = conn.cursor()
    cur.execute(
        "INSERT INTO checklist_items (nombre, orden, activo, tipo_trabajo) VALUES (?, ?, 1, ?)",
        (nombre, orden_i, tipo),
    )
    conn.commit()
    conn.close()
    return RedirectResponse(url=f"/checklist_admin?tipo={tipo}", status_code=303)


@app.post("/checklist_admin/{item_id}/update")
async def checklist_admin_update(
    item_id: int,
    request: Request,
    user=Depends(require_roles("admin")),
):
    form = await request.form()
    nombre = (form.get("nombre") or "").strip()
    tipo = (form.get("tipo_trabajo") or "REPARACION").strip().upper()
    orden = form.get("orden") or "0"
    try:
        orden_i = int(str(orden).strip() or "0")
    except Exception:
        orden_i = 0

    if tipo not in ("REPARACION", "TRAZABILIDAD", "OPTICA_RIGIDA"):
        tipo = "REPARACION"

    conn = get_conn()
    cur = conn.cursor()
    cur.execute(
        "UPDATE checklist_items SET nombre=?, orden=?, tipo_trabajo=? WHERE id=?",
        (nombre, orden_i, tipo, item_id),
    )
    conn.commit()
    conn.close()
    return RedirectResponse(url=f"/checklist_admin?tipo={tipo}", status_code=303)

@app.post("/checklist_admin/{item_id}/toggle")
async def checklist_admin_toggle(item_id: int, request: Request, user=Depends(require_roles("admin"))):
    form = await request.form()
    tipo = (form.get("tipo") or "REPARACION").strip().upper()
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("UPDATE checklist_items SET activo = 1 - COALESCE(activo, 1) WHERE id=?", (item_id,))
    conn.commit()
    conn.close()
    return RedirectResponse(url=f"/checklist_admin?tipo={tipo}", status_code=303)

@app.post("/checklist_admin/{item_id}/delete")
async def checklist_admin_delete(item_id: int, request: Request, user=Depends(require_roles("admin"))):
    form = await request.form()
    tipo = (form.get("tipo") or "REPARACION").strip().upper()
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("DELETE FROM checklist_items WHERE id=?", (item_id,))
    conn.commit()
    conn.close()
    return RedirectResponse(url=f"/checklist_admin?tipo={tipo}", status_code=303)


# -----------------------------
# CATÁLOGO DE REPUESTOS (admin)
# -----------------------------
@app.get("/repuestos_catalogo", response_class=HTMLResponse)
def view_repuestos_catalogo(request: Request, user=Depends(require_roles("admin"))):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("SELECT * FROM repuestos_catalogo ORDER BY nombre")
    items = [dict(r) for r in cur.fetchall()]
    conn.close()
    return templates.TemplateResponse("repuestos_catalogo.html", {"request": request, "user": user, "items": items})

@app.post("/repuestos_catalogo/add")
async def add_repuesto_catalogo(request: Request, user=Depends(require_roles("admin"))):
    form = await request.form()
    nombre = (form.get("nombre") or "").strip()
    precio = float(form.get("precio") or 0)
    if not nombre:
        return RedirectResponse(url="/repuestos_catalogo", status_code=303)
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("INSERT INTO repuestos_catalogo (nombre, precio) VALUES (?, ?)", (nombre, precio))
    conn.commit()
    conn.close()
    return RedirectResponse(url="/repuestos_catalogo", status_code=303)

@app.post("/repuestos_catalogo/{item_id}/update")
async def update_repuesto_catalogo(item_id: int, request: Request, user=Depends(require_roles("admin"))):
    form = await request.form()
    nombre = (form.get("nombre") or "").strip()
    precio = float(form.get("precio") or 0)
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("UPDATE repuestos_catalogo SET nombre=?, precio=? WHERE id=?", (nombre, precio, item_id))
    conn.commit()
    conn.close()
    return RedirectResponse(url="/repuestos_catalogo", status_code=303)

@app.post("/repuestos_catalogo/{item_id}/toggle")
def toggle_repuesto_catalogo(item_id: int, user=Depends(require_roles("admin"))):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("UPDATE repuestos_catalogo SET activo = 1 - COALESCE(activo, 1) WHERE id=?", (item_id,))
    conn.commit()
    conn.close()
    return RedirectResponse(url="/repuestos_catalogo", status_code=303)

@app.post("/repuestos_catalogo/{item_id}/delete")
def delete_repuesto_catalogo(item_id: int, user=Depends(require_roles("admin"))):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("DELETE FROM repuestos_catalogo WHERE id=?", (item_id,))
    conn.commit()
    conn.close()
    return RedirectResponse(url="/repuestos_catalogo", status_code=303)

@app.post("/instrumentos/{instrumento_id}/repuesto/{repuesto_id}/adjust")
def adjust_instrumento_repuesto(instrumento_id: int, repuesto_id: int, action: str = "add", user=Depends(require_roles("admin", "tecnico"))):
    conn = get_conn()
    cur = conn.cursor()
    
    cur.execute("SELECT cantidad FROM instrumento_repuestos WHERE instrumento_id=? AND repuesto_id=?", (instrumento_id, repuesto_id))
    row = cur.fetchone()
    
    if action == "add":
        if row:
            cur.execute("UPDATE instrumento_repuestos SET cantidad = cantidad + 1 WHERE instrumento_id=? AND repuesto_id=?", (instrumento_id, repuesto_id))
        else:
            cur.execute("SELECT precio FROM repuestos_catalogo WHERE id=?", (repuesto_id,))
            rep = cur.fetchone()
            precio = rep["precio"] if rep else 0
            cur.execute("INSERT INTO instrumento_repuestos (instrumento_id, repuesto_id, precio_aplicado, cantidad) VALUES (?, ?, ?, 1)", (instrumento_id, repuesto_id, precio))
    elif action == "sub":
        if row:
            cantidad = row["cantidad"]
            if cantidad > 1:
                cur.execute("UPDATE instrumento_repuestos SET cantidad = cantidad - 1 WHERE instrumento_id=? AND repuesto_id=?", (instrumento_id, repuesto_id))
            else:
                cur.execute("DELETE FROM instrumento_repuestos WHERE instrumento_id=? AND repuesto_id=?", (instrumento_id, repuesto_id))

    # Recalculate total price and info (unit price * count)
    is_pg = os.environ.get("DATABASE_URL") is not None
    # For info, we want something like "Lente x2, Fibra"
    agg_info = "string_agg(CASE WHEN cantidad > 1 THEN nombre || ' x' || cantidad ELSE nombre END, ', ')" if is_pg else \
               "GROUP_CONCAT(CASE WHEN cantidad > 1 THEN nombre || ' x' || cantidad ELSE nombre END, ', ')"
    
    cur.execute(f"""
        SELECT SUM(precio_aplicado * cantidad) as total, {agg_info} as nombres
        FROM instrumento_repuestos ir
        JOIN repuestos_catalogo rc ON ir.repuesto_id = rc.id
        WHERE ir.instrumento_id = ?
    """, (instrumento_id,))
    totals = cur.fetchone()
    
    total_precio = totals["total"] or 0
    total_info = totals["nombres"] or ""
    
    cur.execute("UPDATE instrumentos SET repuesto_precio=?, repuesto_info=? WHERE id=?", (total_precio, total_info, instrumento_id))
    
    # Get current item count for UI feedback
    cur.execute("SELECT cantidad FROM instrumento_repuestos WHERE instrumento_id=? AND repuesto_id=?", (instrumento_id, repuesto_id))
    new_row = cur.fetchone()
    new_qty = new_row["cantidad"] if new_row else 0

    conn.commit()
    conn.close()
    
    return {
        "ok": True, 
        "total_precio": total_precio, 
        "total_info": total_info,
        "cantidad": new_qty
    }


@app.post("/instrumentos/{instrumento_id}/foto_borrar/{slot}")
def foto_borrar(instrumento_id: int, slot: int, user=Depends(require_roles("admin", "recepcion", "tecnico", "grabado"))):
    if slot not in range(1, 7):
        return JSONResponse({"ok": False, "error": "slot inválido"}, status_code=400)

    col = f"foto_entrada_{slot}"

    conn = get_conn()
    cur = conn.cursor()
    cur.execute(f"SELECT {col} FROM instrumentos WHERE id=?", (instrumento_id,))
    row = cur.fetchone()
    if not row:
        conn.close()
        return JSONResponse({"ok": False, "error": "Instrumento no encontrado"}, status_code=404)

    old_path = row[col]
    cur.execute(f"UPDATE instrumentos SET {col}=NULL WHERE id=?", (instrumento_id,))
    conn.commit()
    conn.close()

    _try_delete_public_photo(old_path)
    return {"ok": True}


# -----------------------------
# BORRAR INSTRUMENTO (admin/recepcion)
# -----------------------------
@app.post("/instrumentos/{instrumento_id}/borrar")
def borrar_instrumento(instrumento_id: int, user=Depends(require_roles("admin", "recepcion"))):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("SELECT envio_id, " + ", ".join([f"foto_entrada_{i}" for i in range(1,7)]) + " FROM instrumentos WHERE id=?", (instrumento_id,))
    row = cur.fetchone()
    if not row:
        conn.close()
        return HTMLResponse("Instrumento no encontrado", status_code=404)

    envio_id = row["envio_id"]
    fotos_to_delete = [row[f"foto_entrada_{i}"] for i in range(1,7)]

    cur.execute("DELETE FROM instrumentos WHERE id=?", (instrumento_id,))
    conn.commit()
    conn.close()

    for f in fotos_to_delete:
        _try_delete_public_photo(f)

    return RedirectResponse(url=f"/envios/{envio_id}", status_code=303)


# IMPORTAR EXCEL (admin/recepcion)
# -----------------------------
@app.post("/envios/{envio_id}/importar")
async def envio_importar_excel(
    envio_id: int,
    file: UploadFile = File(...),
    user=Depends(require_roles("admin", "recepcion")),
):
    """Importa instrumentos desde un Excel a un parte existente."""
    path = os.path.join(UPLOAD_DIR, file.filename)
    with open(path, "wb") as f:
        f.write(await file.read())

    # Reutilizamos la lógica de lectura existente
    try:
        _, _, df = leer_excel_envio(path)
    except Exception as e:
        return RedirectResponse(url=f"/envios/{envio_id}?err=excel&msg={str(e)}", status_code=303)

    conn = get_conn()
    cur = conn.cursor()

    rows = []
    now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    for _, r in df.iterrows():
        # Normalizamos un poco los campos
        codigo = str(r.get("codigo_producto") or "").strip()
        if not codigo: continue # si no hay código, saltamos? o permitimos?

        rows.append((
            envio_id,
            codigo,
            str(r.get("fabricante") or "").strip(),
            str(r.get("num_serie") or "").strip(),
            str(r.get("denominacion") or "").strip(),
            str(r.get("observaciones") or "").strip(),
            str(r.get("codigo_datamatrix") or "").strip(),
            str(r.get("codigo_cliente") or "").strip(),
            "", # nombre_trazabilidad
            "Pendiente",
            now_str,
        ))

    if rows:
        cur.executemany("""
            INSERT INTO instrumentos
            (envio_id, codigo_producto, fabricante, num_serie, denominacion, observaciones, codigo_datamatrix, codigo_cliente, nombre_trazabilidad, estado, creado_en)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, rows)
        conn.commit()

    conn.close()
    return RedirectResponse(url=f"/envios/{envio_id}?ok=import", status_code=303)


# -----------------------------
# IMPORTAR EXCEL (admin/recepcion) - CREANDO NUEVA OT
# -----------------------------
@app.get("/importar", response_class=HTMLResponse)
def importar_form(request: Request, user=Depends(require_roles("admin", "recepcion"))):
    conn = get_conn()
    cur = conn.cursor()
    clientes = _list_clientes(cur)
    conn.close()
    return templates.TemplateResponse("importar.html", {"request": request, "user": user, "clientes": clientes})


@app.post("/importar")
async def importar_excel(
    tipo_trabajo: str = Form("REPARACION"),
    referencia: str = Form(""),
    cliente_id: str = Form(""),
    cliente_manual: str = Form("", alias="cliente"),
    fecha_manual: str = Form("", alias="fecha"),
    observaciones: str = Form(""),
    file: UploadFile = File(...),
    user=Depends(require_roles("admin", "recepcion")),
):
    path = os.path.join(UPLOAD_DIR, file.filename)
    with open(path, "wb") as f:
        f.write(await file.read())

    try:
        cliente_auto, fecha_auto, df = leer_excel_envio(path)
    except Exception as e:
        return RedirectResponse(url=f"/importar?err=excel&msg={str(e)}", status_code=303)

    conn = get_conn()
    cur = conn.cursor()

    # Procesar cliente
    final_cliente = ""
    target_cliente_id = None
    
    if cliente_id:
        cur.execute("SELECT nombre FROM clientes WHERE id=?", (cliente_id,))
        row = cur.fetchone()
        if row:
            final_cliente = row["nombre"]
            target_cliente_id = int(cliente_id)
    elif cliente_manual:
        final_cliente = cliente_manual.strip()
    else:
        final_cliente = str(cliente_auto).strip()

    if not final_cliente:
        conn.close()
        return RedirectResponse(url="/importar?err=cliente", status_code=303)

    # Procesar fecha
    final_fecha = fecha_manual if fecha_manual else fecha_auto

    # Procesar nombre archivo / referencia
    final_ref = referencia.strip() if referencia.strip() else file.filename

    ot_num = _next_ot_num(cur)
    tipo_trabajo = (tipo_trabajo or "REPARACION").strip().upper()
    if tipo_trabajo not in ("REPARACION", "TRAZABILIDAD", "OPTICA_RIGIDA"):
        tipo_trabajo = "REPARACION"

    cols = ["ot_num", "nombre_archivo", "cliente", "fecha"]
    vals = [ot_num, final_ref, final_cliente, final_fecha]

    if _envios_has_column(cur, "tipo_trabajo"):
        cols.append("tipo_trabajo")
        vals.append(tipo_trabajo)
    
    if _envios_has_column(cur, "cliente_id") and target_cliente_id:
        cols.append("cliente_id")
        vals.append(target_cliente_id)
        
    if _envios_has_column(cur, "observaciones"):
        cols.append("observaciones")
        vals.append(observaciones)

    qs = ", ".join(["?"] * len(vals))
    sql = f"INSERT INTO envios ({', '.join(cols)}) VALUES ({qs})"
    
    is_pg = os.environ.get("DATABASE_URL") is not None
    if is_pg:
        sql += " RETURNING id"
        cur.execute(sql, tuple(vals))
        row = cur.fetchone()
        if row:
            envio_id = int(row["id"])
        else:
            # Fallback robusto
            cur.execute("SELECT id FROM envios WHERE ot_num=?", (ot_num,))
            row_f = cur.fetchone()
            envio_id = int(row_f["id"]) if row_f else 0
    else:
        cur.execute(sql, tuple(vals))
        envio_id = cur.lastrowid

    if not envio_id or envio_id == 0:
        conn.close()
        raise Exception("No se pudo crear el envío o recuperar su ID. Abortando importación de instrumentos.")

    rows = []
    now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    for _, r in df.iterrows():
        rows.append((
            envio_id,
            r.get("codigo_producto"),
            r.get("fabricante"),
            r.get("num_serie"),
            r.get("denominacion"),
            r.get("observaciones"),
            str(r.get("codigo_datamatrix") or "").strip(),
            str(r.get("codigo_cliente") or "").strip(),
            "", # nombre_trazabilidad
            "Pendiente",
            now_str,
        ))

    cur.executemany("""
        INSERT INTO instrumentos
        (envio_id, codigo_producto, fabricante, num_serie, denominacion, observaciones, codigo_datamatrix, codigo_cliente, nombre_trazabilidad, estado, creado_en)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, rows)

    conn.commit()
    conn.close()
    return RedirectResponse(url=f"/envios/{envio_id}", status_code=303)


# -----------------------------
# USUARIOS (admin)
# -----------------------------
@app.get("/usuarios", response_class=HTMLResponse)
def usuarios_list(request: Request, user=Depends(require_roles("admin"))):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("SELECT id, username, role, is_active, created_at FROM users ORDER BY id DESC;")
    users = [dict(r) for r in cur.fetchall()]
    conn.close()
    return templates.TemplateResponse("users.html", {"request": request, "user": user, "users": users})


@app.get("/usuarios/nuevo", response_class=HTMLResponse)
def usuarios_new_form(request: Request, user=Depends(require_roles("admin"))):
    conn = get_conn()
    cur = conn.cursor()
    clientes = _list_clientes(cur)
    conn.close()
    return templates.TemplateResponse(
        "user_form.html",
        {"request": request, "user": user, "mode": "new", "u": {"username": "", "role": "tecnico", "is_active": 1, "cliente_id": None}, "clientes": clientes, "error": None},
    )


@app.post("/usuarios/nuevo")
def usuarios_new(
    request: Request,
    username: str = Form(...),
    password: str = Form(...),
    role: str = Form(...),
    is_active: int = Form(1),
    cliente_id: Optional[str] = Form(None),
    user=Depends(require_roles("admin")),
):
    username = (username or "").strip()

    if role not in ("admin", "recepcion", "tecnico", "grabado", "cliente"):
        return templates.TemplateResponse(
            "user_form.html",
            {"request": request, "user": user, "mode": "new", "u": {"username": username, "role": role, "is_active": is_active}, "error": "Rol inválido"},
            status_code=400,
        )

    if len(username) < 3:
        return templates.TemplateResponse(
            "user_form.html",
            {"request": request, "user": user, "mode": "new", "u": {"username": username, "role": role, "is_active": is_active}, "error": "El usuario debe tener al menos 3 caracteres"},
            status_code=400,
        )

    if not password or len(password) < 6:
        return templates.TemplateResponse(
            "user_form.html",
            {"request": request, "user": user, "mode": "new", "u": {"username": username, "role": role, "is_active": is_active}, "error": "La contraseña debe tener al menos 6 caracteres"},
            status_code=400,
        )

    conn = get_conn()
    cur = conn.cursor()

    schema = _users_schema(cur)
    pw_col = schema.get("password_col")
    if not pw_col:
        conn.close()
        return templates.TemplateResponse(
            "user_form.html",
            {"request": request, "user": user, "mode": "new", "u": {"username": username, "role": role, "is_active": is_active}, "error": "Error de esquema DB"},
            status_code=400,
        )

    try:
        # Preparamos el cliente_id
        cli_id_val = None
        if cliente_id:
            try:
                cli_id_val = int(cliente_id)
            except:
                cli_id_val = None

        cols = ["username", pw_col, "role", "cliente_id"]
        vals = [username, hash_password(password), role, cli_id_val]
        
        if schema.get("has_is_active"):
            cols.append("is_active")
            vals.append(int(is_active))
            
        if schema.get("has_created_at"):
            cols.append("created_at")
            vals.append(_now_str())
        elif schema.get("has_created_at_at"):
            cols.append("created_at_at")
            vals.append(_now_str())

        sql = f"INSERT INTO users ({', '.join(cols)}) VALUES ({', '.join(['?']*len(cols))})"
        cur.execute(sql, tuple(vals))
        
        # Insert permissions default
        # (Generic logic copied from dash_users_nuevo, simplified here without RETURNING id logic as we redirect anyway)
        # But wait, we need ID to insert permissions.
        # So we MUST fetch ID.
        
        is_pg = bool(os.environ.get("DATABASE_URL"))
        new_id = None
        if is_pg:
             # On PG we need RETURNING id, but we already executed insert without it above? No wait.
             # We should use RETURNING id in the SAME execute or fetch lastrowid is unreliable on PG.
             # Let's redo the execute properly.
             pass
        
    except Exception:
        pass 
        
    # Re-writing the block properly to match dash_users_nuevo logic completely
    
    try:
        # Preparamos el cliente_id
        cli_id_val = None
        if cliente_id:
            try:
                cli_id_val = int(cliente_id)
            except:
                cli_id_val = None

        cols = ["username", pw_col, "role", "cliente_id"]
        vals = [username, hash_password(password), role, cli_id_val]

        if schema.get("has_is_active"):
            cols.append("is_active")
            vals.append(int(is_active))

        if schema.get("has_created_at"):
            cols.append("created_at")
            vals.append(_now_str())
        elif schema.get("has_created_at_at"):
            cols.append("created_at_at")
            vals.append(_now_str())

        sql = f"INSERT INTO users ({', '.join(cols)}) VALUES ({', '.join(['?']*len(cols))})"
        
        is_pg = bool(os.environ.get("DATABASE_URL"))
        if is_pg:
            sql += " RETURNING id"
            cur.execute(sql, tuple(vals))
            row = cur.fetchone()
            if row:
                new_id = int(row["id"])
            else:
                 raise Exception("No ID returned")
        else:
            cur.execute(sql, tuple(vals))
            new_id = int(cur.lastrowid)

        # Default permissions
        for action, _label in ACTIONS:
            allowed = _default_allowed_by_role(role, action)
            cur.execute(
                """
                INSERT INTO user_permissions (user_id, action, allowed) VALUES (?,?,?)
                ON CONFLICT (user_id, action) DO UPDATE SET allowed=excluded.allowed
                """,
                (new_id, action, int(allowed)),
            )

        conn.commit()
    except sqlite3.IntegrityError:
        conn.close()
        return templates.TemplateResponse(
            "user_form.html",
            {"request": request, "user": user, "mode": "new", "u": {"username": username, "role": role, "is_active": is_active}, "error": "Ese usuario ya existe"},
            status_code=400,
        )
    except Exception as e:
        conn.close()
        print(f"Error creating user: {e}")
        return templates.TemplateResponse(
            "user_form.html",
            {"request": request, "user": user, "mode": "new", "u": {"username": username, "role": role, "is_active": is_active}, "error": "Error interno de base de datos"},
            status_code=400,
        )

    conn.close()
    return RedirectResponse(url="/usuarios", status_code=303)


def _users_schema(cur) -> dict:
    from db import get_table_columns
    cols = get_table_columns(cur, "users")
    colset = set(cols)

    # columna de password
    if "password_hash" in colset:
        password_col = "password_hash"
    elif "password" in colset:
        password_col = "password"
    else:
        password_col = None

    return {
        "cols": colset,
        "password_col": password_col,
        "has_is_active": ("is_active" in colset),
        "has_created_at": ("created_at" in colset),
        "has_created_at_at": ("created_at_at" in colset),
    }

def _select_users_sql(schema: dict) -> str:
    parts = ["id", "username", "role", "cliente_id"]

    if schema.get("has_is_active"):
        parts.append("is_active")
    else:
        parts.append("1 AS is_active")

    if schema.get("has_created_at"):
        parts.append("created_at")
    elif schema.get("has_created_at_at"):
        parts.append("created_at_at AS created_at")
    else:
        parts.append("NULL AS created_at")

    return "SELECT " + ", ".join(parts) + " FROM users ORDER BY id ASC"

def _now_str() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")

# -----------------------------
# USUARIOS desde MODAL en DASHBOARD (solo admin)
# No toca tu gestor /usuarios existente.
# -----------------------------
@app.post("/dash_users/nuevo")
def dash_users_nuevo(
    username: str = Form(...),
    password: str = Form(...),
    role: str = Form(...),
    cliente_id: Optional[str] = Form(None),
    user=Depends(require_roles("admin")),
):
    username = (username or "").strip()
    if not username:
        return RedirectResponse(url="/?users=1&uerr=username", status_code=303)

    if role not in ("admin", "recepcion", "tecnico", "grabado", "cliente"):
        return RedirectResponse(url="/?users=1&uerr=role", status_code=303)

    if not (password or "").strip():
        return RedirectResponse(url="/?users=1&uerr=pw", status_code=303)

    conn = get_conn()
    cur = conn.cursor()

    cur.execute("SELECT id FROM users WHERE username=?", (username,))
    if cur.fetchone():
        conn.close()
        return RedirectResponse(url="/?users=1&uerr=exists", status_code=303)

    schema = _users_schema(cur)
    pw_col = schema.get("password_col")
    if not pw_col:
        conn.close()
        return RedirectResponse(url="/?users=1&uerr=db", status_code=303)

    try:
        # Preparamos el cliente_id
        cli_id_val = None
        if cliente_id:
            try:
                cli_id_val = int(cliente_id)
            except:
                cli_id_val = None

        cols = ["username", pw_col, "role", "cliente_id"]
        vals = [username, hash_password(password), role, cli_id_val]

        if schema.get("has_is_active"):
            cols.append("is_active")
            vals.append(1)

        if schema.get("has_created_at"):
            cols.append("created_at")
            vals.append(_now_str())
        elif schema.get("has_created_at_at"):
            cols.append("created_at_at")
            vals.append(_now_str())

        sql = f"INSERT INTO users ({', '.join(cols)}) VALUES ({', '.join(['?']*len(cols))})"
        
        # Postgres necesita RETURNING id porque cursor.lastrowid no es estándar
        is_pg = bool(os.environ.get("DATABASE_URL"))
        if is_pg:
            sql += " RETURNING id"
            cur.execute(sql, tuple(vals))
            row = cur.fetchone()
            if not row:
                raise Exception("No se obtuvo ID del nuevo usuario (Postgres)")
            new_id = int(row["id"])
        else:
            cur.execute(sql, tuple(vals))
            new_id = int(cur.lastrowid)

        for action, _label in ACTIONS:
            allowed = _default_allowed_by_role(role, action)
            cur.execute(
                """
                INSERT INTO user_permissions (user_id, action, allowed) VALUES (?,?,?)
                ON CONFLICT (user_id, action) DO UPDATE SET allowed=excluded.allowed
                """,
                (new_id, action, int(allowed)),
            )

        conn.commit()
    except Exception as e:
        print(f"ERROR creando usuario: {e}")
        conn.close()
        return RedirectResponse(url="/?users=1&uerr=db", status_code=303)

    conn.close()
    return RedirectResponse(url="/?users=1&uok=created", status_code=303)



@app.post("/dash_users/{user_id}/role")
def dash_users_set_role(
    user_id: int,
    role: str = Form(...),
    user=Depends(require_roles("admin")),
):
    if role not in ("admin", "recepcion", "tecnico", "grabado", "cliente"):
        return RedirectResponse(url="/?users=1&uerr=role", status_code=303)

    conn = get_conn()
    cur = conn.cursor()
    cur.execute("UPDATE users SET role=? WHERE id=?", (role, user_id))

    # Ajusta permisos a defaults del rol (sin borrar nada extra)
    for action, _label in ACTIONS:
        allowed = _default_allowed_by_role(role, action)
        cur.execute(
            """
            INSERT INTO user_permissions (user_id, action, allowed) VALUES (?,?,?)
            ON CONFLICT (user_id, action) DO UPDATE SET allowed=excluded.allowed
            """,
            (int(user_id), action, int(allowed)),
        )

    conn.commit()
    conn.close()
    return RedirectResponse(url="/?users=1&uok=role", status_code=303)


@app.post("/dash_users/{user_id}/cliente")
def dash_users_set_cliente(
    user_id: int,
    cliente_id: Optional[str] = Form(None),
    user=Depends(require_roles("admin")),
):
    cli_id_val = None
    if cliente_id:
        try:
            cli_id_val = int(cliente_id)
        except:
            cli_id_val = None

    conn = get_conn()
    cur = conn.cursor()
    cur.execute("UPDATE users SET cliente_id=? WHERE id=?", (cli_id_val, user_id))
    conn.commit()
    conn.close()
    return RedirectResponse(url="/?users=1&uok=cliente", status_code=303)


@app.post("/dash_users/{user_id}/toggle")
def dash_users_toggle_active(
    user_id: int,
    user=Depends(require_roles("admin")),
):
    conn = get_conn()
    cur = conn.cursor()
    schema = _users_schema(cur)
    if not schema.get("has_is_active"):
        conn.close()
        return RedirectResponse(url="/?users=1&uerr=db", status_code=303)

    cur.execute("SELECT is_active FROM users WHERE id=?", (user_id,))
    row = cur.fetchone()
    if not row:
        conn.close()
        return RedirectResponse(url="/?users=1&uerr=notfound", status_code=303)

    new_val = 0 if int(row["is_active"] or 0) == 1 else 1
    cur.execute("UPDATE users SET is_active=? WHERE id=?", (new_val, user_id))
    conn.commit()
    conn.close()
    return RedirectResponse(url="/?users=1&uok=active", status_code=303)



@app.post("/dash_users/{user_id}/password")
def dash_users_set_password(
    user_id: int,
    password: str = Form(...),
    user=Depends(require_roles("admin")),
):
    if not (password or "").strip():
        return RedirectResponse(url="/?users=1&uerr=pw", status_code=303)

    conn = get_conn()
    cur = conn.cursor()
    schema = _users_schema(cur)
    pw_col = schema.get("password_col")
    if not pw_col:
        conn.close()
        return RedirectResponse(url="/?users=1&uerr=db", status_code=303)

    cur.execute(f"UPDATE users SET {pw_col}=? WHERE id=?", (hash_password(password), user_id))
    conn.commit()
    conn.close()
    return RedirectResponse(url="/?users=1&uok=pw", status_code=303)



@app.post("/dash_users/{user_id}/perms")
async def dash_users_set_perms(
    request: Request,
    user_id: int,
    user=Depends(require_roles("admin")),
):
    form = await request.form()

    conn = get_conn()
    cur = conn.cursor()
    cur.execute("SELECT id FROM users WHERE id=?", (user_id,))
    if not cur.fetchone():
        conn.close()
        return RedirectResponse(url="/?users=1&uerr=notfound", status_code=303)

    for action, _label in ACTIONS:
        allowed = 1 if form.get(f"perm_{action}") == "on" else 0
        cur.execute(
            """
            INSERT INTO user_permissions (user_id, action, allowed) VALUES (?,?,?)
            ON CONFLICT (user_id, action) DO UPDATE SET allowed=excluded.allowed
            """,
            (int(user_id), action, int(allowed)),
        )

    conn.commit()
    conn.close()
    return RedirectResponse(url="/?users=1&uok=perms", status_code=303)

@app.post("/dash_users/{user_id}/delete")
def dash_users_delete(
    user_id: int,
    user=Depends(require_roles("admin")),
):
    # Evitar que el admin se borre a sí mismo
    if _user_id(user) == int(user_id):
        return RedirectResponse(url="/?users=1&uerr=selfdelete", status_code=303)

    conn = get_conn()
    cur = conn.cursor()

    cur.execute("SELECT id FROM users WHERE id=?", (user_id,))
    if not cur.fetchone():
        conn.close()
        return RedirectResponse(url="/?users=1&uerr=notfound", status_code=303)

    cur.execute("DELETE FROM user_permissions WHERE user_id=?", (user_id,))
    cur.execute("DELETE FROM users WHERE id=?", (user_id,))

    conn.commit()
    conn.close()
    return RedirectResponse(url="/?users=1&uok=deleted", status_code=303)


# -----------------------------
# BORRAR PARTE / ENVÍO (admin/recepcion)


# -----------------------------
# USUARIOS desde MODAL en DASHBOARD (solo admin)
# No toca tu gestor /usuarios existente.
# -----------------------------




@app.post("/dash_users/{user_id}/role")
def dash_users_set_role(
    user_id: int,
    role: str = Form(...),
    user=Depends(require_roles("admin")),
):
    if role not in ("admin", "recepcion", "tecnico", "grabado"):
        return RedirectResponse(url="/?users=1&uerr=role", status_code=303)

    conn = get_conn()
    cur = conn.cursor()
    cur.execute("UPDATE users SET role=? WHERE id=?", (role, user_id))

    # Ajusta permisos a defaults del rol (sin borrar nada extra)
    for action, _label in ACTIONS:
        allowed = _default_allowed_by_role(role, action)
        cur.execute(
            "INSERT OR REPLACE INTO user_permissions (user_id, action, allowed) VALUES (?,?,?)",
            (int(user_id), action, int(allowed)),
        )

    conn.commit()
    conn.close()
    return RedirectResponse(url="/?users=1&uok=role", status_code=303)


@app.post("/dash_users/{user_id}/toggle")
def dash_users_toggle_active(
    user_id: int,
    user=Depends(require_roles("admin")),
):
    conn = get_conn()
    cur = conn.cursor()
    schema = _users_schema(cur)
    if not schema.get("has_is_active"):
        conn.close()
        return RedirectResponse(url="/?users=1&uerr=db", status_code=303)

    cur.execute("SELECT is_active FROM users WHERE id=?", (user_id,))
    row = cur.fetchone()
    if not row:
        conn.close()
        return RedirectResponse(url="/?users=1&uerr=notfound", status_code=303)

    new_val = 0 if int(row["is_active"] or 0) == 1 else 1
    cur.execute("UPDATE users SET is_active=? WHERE id=?", (new_val, user_id))
    conn.commit()
    conn.close()
    return RedirectResponse(url="/?users=1&uok=active", status_code=303)



@app.post("/dash_users/{user_id}/password")
def dash_users_set_password(
    user_id: int,
    password: str = Form(...),
    user=Depends(require_roles("admin")),
):
    if not (password or "").strip():
        return RedirectResponse(url="/?users=1&uerr=pw", status_code=303)

    conn = get_conn()
    cur = conn.cursor()
    schema = _users_schema(cur)
    pw_col = schema.get("password_col")
    if not pw_col:
        conn.close()
        return RedirectResponse(url="/?users=1&uerr=db", status_code=303)

    cur.execute(f"UPDATE users SET {pw_col}=? WHERE id=?", (hash_password(password), user_id))
    conn.commit()
    conn.close()
    return RedirectResponse(url="/?users=1&uok=pw", status_code=303)



@app.post("/dash_users/{user_id}/perms")
async def dash_users_set_perms(
    request: Request,
    user_id: int,
    user=Depends(require_roles("admin")),
):
    form = await request.form()

    conn = get_conn()
    cur = conn.cursor()
    cur.execute("SELECT id FROM users WHERE id=?", (user_id,))
    if not cur.fetchone():
        conn.close()
        return RedirectResponse(url="/?users=1&uerr=notfound", status_code=303)

    for action, _label in ACTIONS:
        allowed = 1 if form.get(f"perm_{action}") == "on" else 0
        cur.execute(
            "INSERT OR REPLACE INTO user_permissions (user_id, action, allowed) VALUES (?,?,?)",
            (int(user_id), action, int(allowed)),
        )

    conn.commit()
    conn.close()
    return RedirectResponse(url="/?users=1&uok=perms", status_code=303)

# -----------------------------
# BORRAR PARTE / ENVÍO (admin/recepcion)
#  - Bloquea borrado si el parte está "cerrado" (sin pendientes y con instrumentos)
#  - Requiere confirmación fuerte: escribir la OT exacta
# -----------------------------
@app.post("/envios/{envio_id}/borrar")
def borrar_envio(
    request: Request,
    envio_id: int,
    confirm_ot: str = Form(""),
    user=Depends(require_roles("admin", "recepcion")),
):
    # Permiso granular (además del rol)
    conn_perm = get_conn()
    cur_perm = conn_perm.cursor()
    if not can_action(user, "envio_borrar", cur_perm):
        conn_perm.close()
        return RedirectResponse(url="/?err=perm", status_code=303)
    conn_perm.close()

    conn = get_conn()
    cur = conn.cursor()

    cur.execute("SELECT id, ot_num FROM envios WHERE id=?", (envio_id,))
    envio = cur.fetchone()
    if not envio:
        conn.close()
        return RedirectResponse(url="/?err=envio_no_encontrado", status_code=303)

    ot_num = (envio["ot_num"] or "").strip()

    # Confirmación fuerte
    if (confirm_ot or "").strip() != ot_num:
        conn.close()
        return RedirectResponse(url="/?err=confirm_ot", status_code=303)

    # Bloquear borrado si está cerrado:
    # "cerrado" = tiene instrumentos y ninguno está en Pendiente/En proceso
    cur.execute("SELECT COUNT(*) AS n FROM instrumentos WHERE envio_id=?", (envio_id,))
    n_inst = int(cur.fetchone()["n"] or 0)

    cur.execute(
        "SELECT COUNT(*) AS n FROM instrumentos WHERE envio_id=? AND estado IN ('Pendiente','En proceso')",
        (envio_id,),
    )
    n_pend = int(cur.fetchone()["n"] or 0)

    if n_inst > 0 and n_pend == 0:
        conn.close()
        return RedirectResponse(url="/?err=cerrado", status_code=303)

    # Borrar fotos de instrumentos del envío (evita archivos huérfanos)
    cur.execute(
        "SELECT foto_entrada_1, foto_entrada_2 FROM instrumentos WHERE envio_id=?",
        (envio_id,)
    )
    for r in cur.fetchall():
        _try_delete_public_photo(r["foto_entrada_1"])
        _try_delete_public_photo(r["foto_entrada_2"])

    # Borrar instrumentos y envío
    cur.execute("DELETE FROM instrumentos WHERE envio_id=?", (envio_id,))
    cur.execute("DELETE FROM envios WHERE id=?", (envio_id,))

    conn.commit()
    conn.close()

    return RedirectResponse(url="/?ok=borrado", status_code=303)


@app.post("/peticion_recogida")
async def peticion_recogida(
    n_instrumentos: int = Form(0),
    contacto: str = Form(""),
    telefono: str = Form(""),
    observaciones: str = Form(""),
    user=Depends(require_roles("cliente", "admin", "recepcion"))
):
    """El cliente solicita que se pase a recoger material."""
    cli_id = (user.get("cliente_id") if isinstance(user, dict) else getattr(user, "cliente_id", None))
    u_id = (user.get("id") if isinstance(user, dict) else getattr(user, "id", None))

    if not cli_id:
        return RedirectResponse(url="/?err=nocli", status_code=303)

    conn = get_conn()
    cur = conn.cursor()

    # Generar número de petición: REC-YYMM-XXX
    now = datetime.now()
    prefix = f"REC-{now.strftime('%y%m')}-"
    
    # Buscar el último número del mes actual
    cur.execute("SELECT num_peticion FROM peticiones_recogida WHERE num_peticion LIKE ? ORDER BY id DESC LIMIT 1", (f"{prefix}%",))
    row = cur.fetchone()
    if row and row["num_peticion"]:
        try:
            last_num = int(row["num_peticion"].split("-")[-1])
            new_num = last_num + 1
        except:
            new_num = 1
    else:
        new_num = 1
    
    num_peticion = f"{prefix}{str(new_num).zfill(3)}"

    cur.execute("""
        INSERT INTO peticiones_recogida (num_peticion, cliente_id, usuario_id, n_instrumentos, contacto, telefono, observaciones, estado, creado_en)
        VALUES (?, ?, ?, ?, ?, ?, ?, 'Pendiente', CURRENT_TIMESTAMP)
    """, (num_peticion, int(cli_id or 0), u_id, n_instrumentos, contacto, telefono, observaciones))
    conn.commit()
    conn.close()

    return RedirectResponse(url=f"/?msg=recogida_ok&num={num_peticion}", status_code=303)

@app.get("/recogidas", response_class=HTMLResponse)
def recogidas_list(request: Request, user=Depends(get_current_user)):
    role = _user_role(user)
    if role not in ["admin", "recepcion", "cliente"]:
        return RedirectResponse(url="/?err=perm", status_code=303)

    conn = get_conn()
    cur = conn.cursor()
    
    where_sql = ""
    params = []
    if role == "cliente":
        cli_id = (user.get("cliente_id") if isinstance(user, dict) else getattr(user, "cliente_id", None))
        where_sql = "WHERE pr.cliente_id = ?"
        params = [int(cli_id or 0)]

    cur.execute(f"""
        SELECT pr.*, c.nombre as cliente_nombre 
        FROM peticiones_recogida pr
        JOIN clientes c ON pr.cliente_id = c.id
        {where_sql}
        ORDER BY pr.creado_en DESC
    """, tuple(params))
    items = [dict(r) for r in cur.fetchall()]
    conn.close()
    return templates.TemplateResponse("recogidas_list.html", {"request": request, "user": user, "items": items})


@app.post("/peticion_recogida/{peticion_id}/completar")
def completar_peticion_recogida(peticion_id: int, user=Depends(require_roles("admin", "recepcion"))):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("UPDATE peticiones_recogida SET estado = 'Completada' WHERE id = ?", (peticion_id,))
    conn.commit()
    conn.close()
    return RedirectResponse(url="/", status_code=303)


@app.get("/api/peticiones_recogida/count")
def count_peticiones_recogida(user=Depends(require_roles("admin", "recepcion"))):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("SELECT COUNT(*) as n FROM peticiones_recogida WHERE estado = 'Pendiente'")
    row = cur.fetchone()
    conn.close()
    return {"count": int(row["n"] or 0) if row else 0}

@app.get("/api/consultas/unread_count")
def count_consultas_unread(user=Depends(get_current_user)):
    conn = get_conn()
    cur = conn.cursor()
    role = _user_role(user)
    count = 0
    if role in ["admin", "recepcion", "tecnico"]:
        cur.execute("SELECT COUNT(*) as n FROM consultas WHERE estado='Abierta'")
        count = cur.fetchone()["n"]
    elif role == "cliente" and user.get("cliente_id"):
        cur.execute("SELECT COUNT(*) as n FROM consultas WHERE cliente_id=? AND estado='Respondida'", (int(user.get("cliente_id")),))
        count = cur.fetchone()["n"]
    conn.close()
    return {"count": int(count or 0)}


# -----------------------------
# CONSULTAS TÉCNICAS (Chat)
# -----------------------------
@app.post("/consultas/nueva")
async def consulta_nueva(
    titulo: str = Form(...),
    descripcion: str = Form(...),
    foto1: UploadFile = File(None),
    foto2: UploadFile = File(None),
    foto3: UploadFile = File(None),
    user=Depends(get_current_user)
):
    u_id = (user.get("id") if isinstance(user, dict) else getattr(user, "id", None))
    cli_id = (user.get("cliente_id") if isinstance(user, dict) else getattr(user, "cliente_id", None))
    
    if not cli_id and _user_role(user) == "cliente":
        return RedirectResponse(url="/?err=nocli", status_code=303)
        
    fotos_urls = []
    for f in [foto1, foto2, foto3]:
        if f and f.filename:
            safe_name = re.sub(r"[^a-zA-Z0-9.-]", "_", f.filename)
            fname = f"consulta_{int(time.time())}_{uuid.uuid4().hex[:6]}_{safe_name}"
            path = os.path.join(FOTOS_DIR, fname)
            content = await f.read()
            with open(path, "wb") as out:
                out.write(content)
            fotos_urls.append(f"/static/fotos/{fname}")
        else:
            fotos_urls.append(None)

    conn = get_conn()
    cur = conn.cursor()
    cur.execute("""
        INSERT INTO consultas (cliente_id, usuario_id, titulo, descripcion, foto_1, foto_2, foto_3)
        VALUES (?, ?, ?, ?, ?, ?, ?)
    """, (cli_id or 0, u_id, titulo, descripcion, fotos_urls[0], fotos_urls[1], fotos_urls[2]))
    conn.commit()
    conn.close()
    
    return RedirectResponse(url="/?msg=consulta_ok", status_code=303)

@app.get("/consultas/{consulta_id}", response_class=HTMLResponse)
async def consulta_detalle(request: Request, consulta_id: int, user=Depends(get_current_user)):
    conn = get_conn()
    cur = conn.cursor()
    
    cur.execute("""
        SELECT c.*, cl.nombre as cliente_nombre 
        FROM consultas c 
        LEFT JOIN clientes cl ON c.cliente_id = cl.id 
        WHERE c.id = ?
    """, (consulta_id,))
    consulta = cur.fetchone()
    if not consulta:
        conn.close()
        return HTMLResponse("Consulta no encontrada", status_code=404)
    
    # Seguridad
    if _user_role(user) == "cliente":
        u_cli_id = (user.get("cliente_id") if isinstance(user, dict) else getattr(user, "cliente_id", None))
        if int(consulta.get("cliente_id", 0)) != int(u_cli_id or -1):
            conn.close()
            return HTMLResponse("Acceso denegado", status_code=403)
            
    cur.execute("""
        SELECT m.*, u.username, u.role
        FROM consultas_mensajes m
        JOIN users u ON m.usuario_id = u.id
        WHERE m.consulta_id = ?
        ORDER BY m.creado_en ASC
    """, (consulta_id,))
    mensajes = [dict(r) for r in cur.fetchall()]
    
    conn.close()
    return templates.TemplateResponse("consulta_chat.html", {
        "request": request,
        "user": user,
        "consulta": dict(consulta),
        "mensajes": mensajes
    })

@app.post("/consultas/{consulta_id}/mensaje")
async def consulta_mensaje_enviar(consulta_id: int, mensaje: str = Form(...), user=Depends(get_current_user)):
    u_id = (user.get("id") if isinstance(user, dict) else getattr(user, "id", None))
    
    conn = get_conn()
    cur = conn.cursor()
    
    # Actualizar estado si el que escribe es admin/recepcion/tecnico
    nuevo_estado = "Respondida" if _user_role(user) in ["admin", "recepcion", "tecnico"] else "Abierta"
    
    cur.execute("INSERT INTO consultas_mensajes (consulta_id, usuario_id, mensaje) VALUES (?, ?, ?)", (consulta_id, u_id, mensaje))
    cur.execute("UPDATE consultas SET estado = ?, actualizado_en = CURRENT_TIMESTAMP WHERE id = ?", (nuevo_estado, consulta_id))
    
    conn.commit()
    conn.close()
    return RedirectResponse(url=f"/consultas/{consulta_id}", status_code=303)

@app.post("/consultas/{consulta_id}/cerrar")
def cerrar_consulta(consulta_id: int, user=Depends(require_roles("admin", "recepcion", "tecnico"))):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("UPDATE consultas SET estado = 'Cerrada' WHERE id = ?", (consulta_id,))
    conn.commit()
    conn.close()
    return RedirectResponse(url="/", status_code=303)

@app.get("/estadisticas_tecnicos", response_class=HTMLResponse)
def estadisticas_tecnicos(
    request: Request,
    fecha_inicio: Optional[str] = None,
    fecha_fin: Optional[str] = None,
    user=Depends(require_roles("admin"))
):
    conn = get_conn()
    cur = conn.cursor()

    # Filtros por defecto (mes actual si no se especifica)
    hoy = datetime.now()
    if not fecha_inicio:
        fecha_inicio = hoy.replace(day=1).strftime("%Y-%m-%d")
    if not fecha_fin:
        fecha_fin = hoy.strftime("%Y-%m-%d")

    # SQL para contar por técnico y estado
    # Ajustamos fecha_fin para incluir todo el día (hasta 23:59:59)
    fecha_fin_ts = f"{fecha_fin} 23:59:59"
    fecha_inicio_ts = f"{fecha_inicio} 00:00:00"

    sql = """
        SELECT 
            tecnico_reparacion as tecnico,
            COUNT(CASE WHEN estado = 'Reparado' THEN 1 END) as reparados,
            COUNT(CASE WHEN estado = 'Baja' THEN 1 END) as bajas,
            COUNT(CASE WHEN estado NOT IN ('Reparado', 'Baja') THEN 1 END) as curso,
            COUNT(*) as total
        FROM instrumentos
        WHERE tecnico_reparacion IS NOT NULL 
          AND tecnico_reparacion != ''
          AND tecnico_reparacion_en >= ? 
          AND tecnico_reparacion_en <= ?
        GROUP BY tecnico_reparacion
        ORDER BY reparados DESC
    """

    cur.execute(sql, (fecha_inicio_ts, fecha_fin_ts))
    stats = [dict(r) for r in cur.fetchall()]

    # Totales globales del periodo
    total_reparados = sum(s['reparados'] for s in stats)
    total_bajas = sum(s['bajas'] for s in stats)
    total_curso = sum(s['curso'] for s in stats)
    total_general = sum(s['total'] for s in stats)

    conn.close()

    return templates.TemplateResponse("estadisticas_tecnicos.html", {
        "request": request,
        "user": user,
        "stats": stats,
        "fecha_inicio": fecha_inicio,
        "fecha_fin": fecha_fin,
        "totales": {
            "reparados": total_reparados,
            "bajas": total_bajas,
            "curso": total_curso,
            "total": total_general
        }
    })
