import os
import io
import csv
from datetime import datetime
from typing import Optional, List, Dict, Any

from fastapi import APIRouter, Request, Depends, HTTPException, Form
from fastapi.responses import HTMLResponse, StreamingResponse, RedirectResponse
from fastapi.templating import Jinja2Templates

from db import get_conn
from security import get_current_user, require_roles, hash_password
from shared import (
    _user_role, _user_id, _users_schema, can_action
)
from utils import format_fecha, _norm_codigo, _codigo_variants

try:
    from openpyxl import Workbook
except ImportError:
    Workbook = None

router = APIRouter()
templates = Jinja2Templates(directory="templates")

@router.get("/api/trazabilidad/buscar")
def api_trazabilidad_buscar(q: str = "", user=Depends(get_current_user)):
    q = (q or "").strip()
    if not q: return {"results": []}
    conn = get_conn(); cur = conn.cursor()
    where = ["(i.codigo_datamatrix = ? OR i.num_serie = ? OR i.nombre_trazabilidad = ?)"]
    params = [q, q, q]
    if _user_role(user) == "cliente":
        cli_id = (user.get("cliente_id") if isinstance(user, dict) else getattr(user, "cliente_id", None))
        if cli_id: where.append("e.cliente_id = ?"); params.append(int(cli_id))
        else: conn.close(); return {"results": []}
    cur.execute(f"SELECT i.id, i.envio_id, i.denominacion, i.codigo_producto, i.fabricante, i.estado, i.creado_en, e.ot_num, e.fecha, e.tipo_trabajo FROM instrumentos i JOIN envios e ON e.id = i.envio_id WHERE {' AND '.join(where)} ORDER BY i.creado_en DESC LIMIT 50", tuple(params))
    results = [dict(r) for r in cur.fetchall()]
    for r in results:
        if r.get("fecha"): r["fecha_fmt"] = format_fecha(r["fecha"])
        if r.get("creado_en"): r["creado_fmt"] = format_fecha(r["creado_en"])
    conn.close()
    return {"results": results}

@router.post("/perfil/password")
def change_own_password(password: str = Form(...), user=Depends(get_current_user)):
    if not (password or "").strip() or len(password) < 6: return RedirectResponse(url="/?err=pw_too_short", status_code=303)
    conn = get_conn(); cur = conn.cursor()
    schema = _users_schema(cur); pw_col = schema.get("password_col")
    user_id = _user_id(user)
    if not pw_col or not user_id: conn.close(); return RedirectResponse(url="/?err=db", status_code=303)
    cur.execute(f"UPDATE users SET {pw_col}=? WHERE id=?", (hash_password(password), user_id))
    conn.commit(); conn.close()
    resp = RedirectResponse(url="/login?msg=pw_changed", status_code=303)
    resp.delete_cookie("xurgical_session")
    return resp

@router.get("/export", response_class=HTMLResponse)
def export_home(request: Request, user=Depends(require_roles("admin", "recepcion", "cliente", "socio"))):
    conn = get_conn(); cur = conn.cursor()
    if _user_role(user) == "cliente" and user.get("cliente_id"):
        cur.execute("SELECT id, ot_num, cliente, fecha FROM envios WHERE cliente_id=? ORDER BY id DESC LIMIT 400", (int(user["cliente_id"]),))
    elif _user_role(user) == "cliente": cur.execute("SELECT id, ot_num, cliente, fecha FROM envios WHERE 1=0")
    else: cur.execute("SELECT id, ot_num, cliente, fecha FROM envios ORDER BY id DESC LIMIT 400")
    envios = [dict(r) for r in cur.fetchall()]
    conn.close()
    return templates.TemplateResponse("export.html", {"request": request, "user": user, "envios": envios, "estados": ["TODOS", "Pendiente", "En proceso", "Reparado", "Baja"]})

@router.get("/guias", response_class=HTMLResponse)
def guias_cliente(request: Request, user=Depends(get_current_user)):
    return templates.TemplateResponse("guias.html", {"request": request, "user": user})

@router.get("/export/download")
def export_download(request: Request, user=Depends(require_roles("admin", "recepcion", "cliente", "socio")), scope: str = "partes", envio_id: str = None, estado: str = "TODOS", solo_pendientes: bool = False, grabado: str = "todos", fmt: str = "xlsx"):
    conn = get_conn(); cur = conn.cursor()
    cli_id = (user.get("cliente_id") if isinstance(user, dict) else getattr(user, "cliente_id", None))
    is_cliente = _user_role(user) == "cliente"
    
    partes_rows, inst_rows = [], []
    if scope in {"partes", "parte"}:
        where_env, params_env = "1=1", []
        if envio_id: where_env = "e.id=?"; params_env = [int(envio_id)]
        if is_cliente:
            if cli_id: where_env += " AND e.cliente_id=?"; params_env.append(int(cli_id))
            else: where_env += " AND 1=0"
        cur.execute(f"SELECT e.*, COUNT(i.id) AS n_instrumentos FROM envios e LEFT JOIN instrumentos i ON i.envio_id = e.id WHERE {where_env} GROUP BY e.id ORDER BY e.id DESC", tuple(params_env))
        partes_rows = [dict(r) for r in cur.fetchall()]
    
    if scope in {"instrumentos", "parte"}:
        where_i, params_i = "1=1", []
        if solo_pendientes: where_i = "i.estado IN ('Pendiente','En proceso')"
        elif estado != "TODOS": where_i = "i.estado = ?"; params_i = [estado]
        where_envio, params_envio = "1=1", []
        if scope == "parte": where_envio = "i.envio_id=?"; params_envio = [int(envio_id)]
        if is_cliente:
            if cli_id: where_envio += " AND e.cliente_id=?"; params_envio.append(int(cli_id))
            else: where_envio += " AND 1=0"
        cur.execute(f"SELECT i.*, e.ot_num, e.cliente FROM instrumentos i JOIN envios e ON e.id = i.envio_id WHERE {where_envio} AND {where_i} ORDER BY e.id DESC, i.id ASC", tuple(params_envio + params_i))
        inst_rows = [dict(r) for r in cur.fetchall()]
    conn.close()

    now_tag = datetime.now().strftime("%Y%m%d_%H%M")
    base_name = f"xurgical_export_{scope}_{now_tag}"
    
    if fmt == "csv":
        rows = inst_rows or partes_rows
        headers = list(rows[0].keys()) if rows else []
        sio = io.StringIO(); w = csv.writer(sio, delimiter=";")
        w.writerow(headers)
        for r in rows: w.writerow([r.get(h, "") for h in headers])
        return StreamingResponse(io.BytesIO(sio.getvalue().encode("utf-8-sig")), media_type="text/csv", headers={"Content-Disposition": f"attachment; filename={base_name}.csv"})

    if Workbook is None: raise HTTPException(status_code=500, detail="openpyxl no disponible")
    wb = Workbook(); wb.remove(wb.active)
    if partes_rows:
        ws = wb.create_sheet("Partes"); ws.append(list(partes_rows[0].keys()))
        for r in partes_rows: ws.append(list(r.values()))
    if inst_rows:
        ws = wb.create_sheet("Instrumentos"); ws.append(list(inst_rows[0].keys()))
        for r in inst_rows: ws.append(list(r.values()))
    bio = io.BytesIO(); wb.save(bio); bio.seek(0)
    return StreamingResponse(bio, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"attachment; filename={base_name}.xlsx"})
